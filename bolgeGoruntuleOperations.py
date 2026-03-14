from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QPushButton, QComboBox, QTableWidget, QTableWidgetItem,
                             QHeaderView, QMessageBox, QLineEdit, QDateEdit, QDialog,
                             QDialogButtonBox, QFormLayout, QDoubleSpinBox, QTabWidget,
                             QGroupBox, QGridLayout, QFileDialog)
from PyQt5.QtCore import Qt, QDate
from PyQt5.QtGui import QFont
import requests
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class BolgeGoruntuleTab(QWidget):
    def __init__(self, api_client, user_id, role, bolge_kodlari):
        super().__init__()
        self.api_client = api_client
        self.user_id = user_id
        self.role = role
        self.bolge_kodlari = bolge_kodlari
        
        self.setup_ui()
        self.load_data()
    
    def setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Başlık
        title = QLabel("📈 Bölgelere Göre Görüntüle")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("font-size: 24px; font-weight: bold; margin: 20px;")
        layout.addWidget(title)
        
        # Filtreler - TEST İÇİN KALDIRILDI
        # filter_layout = QHBoxLayout()
        # 
        # # Bölge seçimi
        # filter_layout.addWidget(QLabel("Bölge:"))
        # self.bolge_combo = QComboBox()
        # self.bolge_combo.addItem("Tümü", None)
        # 
        # # Normal kullanıcı ise sadece kendi bölgelerini göster
        # if self.role == 'normal':
        #     for bolge_kodu in self.bolge_kodlari:
        #         self.bolge_combo.addItem(f"Bölge {bolge_kodu}", bolge_kodu)
        # else:
        #     # Admin ve üst düzey yönetici için tüm bölgeler
        #     try:
        #         response = requests.get(f"http://127.0.0.1:5000/api/bolge_kodlari?user_id={self.user_id}")
        #         if response.status_code == 200:
        #             bolge_kodlari_dict = response.json()
        #             for kod, ad in bolge_kodlari_dict.items():
        #                 self.bolge_combo.addItem(f"{ad} ({kod})", kod)
        #     except Exception as e:
        #         print(f"Bölge kodları yüklenirken hata: {e}")
        # 
        # filter_layout.addWidget(self.bolge_combo)
        # 
        # # Safha seçimi
        # filter_layout.addWidget(QLabel("Safha:"))
        # self.safha_combo = QComboBox()
        # self.safha_combo.addItem("Tümü", None)
        # filter_layout.addWidget(self.safha_combo)
        # 
        # # Stage seçimi
        # filter_layout.addWidget(QLabel("Stage:"))
        # self.stage_combo = QComboBox()
        # self.stage_combo.addItem("Tümü", None)
        # filter_layout.addWidget(self.stage_combo)
        # 
        # # Filtrele butonu
        # filter_btn = QPushButton("Filtrele")
        # filter_btn.clicked.connect(self.apply_filters)
        # filter_layout.addWidget(filter_btn)
        # 
        # # Yenile butonu
        # refresh_btn = QPushButton("Yenile")
        # refresh_btn.clicked.connect(self.load_data)
        # filter_layout.addWidget(refresh_btn)
        # 
        # layout.addLayout(filter_layout)
        
        # Test için sadece yenile butonu
        refresh_layout = QHBoxLayout()
        refresh_btn = QPushButton("Yenile")
        refresh_btn.clicked.connect(self.load_data)
        refresh_layout.addWidget(refresh_btn)
        refresh_layout.addStretch()
        layout.addLayout(refresh_layout)
        
        # Combo box'ları None olarak başlat (filtreleme yapılmayacak)
        self.bolge_combo = None
        self.safha_combo = None
        self.stage_combo = None
        
        # Tab widget (Harcama Talep ve Masraf sekmeleri)
        self.tab_widget = QTabWidget()
        
        # Harcama Talep sekmesi
        self.harcama_talep_tab = QWidget()
        self.setup_harcama_talep_tab()
        self.tab_widget.addTab(self.harcama_talep_tab, "Harcama Talep")
        
        # Masraf sekmesi
        self.masraf_tab = QWidget()
        self.setup_masraf_tab()
        self.tab_widget.addTab(self.masraf_tab, "Masraf")
        
        layout.addWidget(self.tab_widget)
    
    def setup_harcama_talep_tab(self):
        layout = QVBoxLayout(self.harcama_talep_tab)
        
        # Tablo
        self.harcama_table = QTableWidget()
        self.harcama_table.setColumnCount(14)  # ID kolonu eklendi
        self.harcama_table.setHorizontalHeaderLabels([
            "ID", "No", "Tarih", "BÖLGE KODU", "KAYNAK TİPİ KODU", "STAGE KODU",
            "STAGE-OPERASYON KODU", "Safha", "Harcama Kalemi", "Birim",
            "Miktar", "Birim ücret", "Toplam", "Açıklama"
        ])
        # ID kolonunu gizle
        self.harcama_table.setColumnHidden(0, True)
        self.harcama_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.harcama_table)
        
        # Analiz verileri (üst düzey yönetici ve normal kullanıcı için)
        if self.role in ['normal', 'ust_duzey_yonetici', 'admin']:
            analysis_group = QGroupBox("📊 Analiz Verileri")
            analysis_layout = QGridLayout()
            
            self.total_label = QLabel("Toplam Tutar: 0 ₺")
            self.total_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2563eb;")
            analysis_layout.addWidget(self.total_label, 0, 0)
            
            self.count_label = QLabel("Toplam Kayıt: 0")
            self.count_label.setStyleSheet("font-size: 14px;")
            analysis_layout.addWidget(self.count_label, 0, 1)
            
            self.avg_label = QLabel("Ortalama Tutar: 0 ₺")
            self.avg_label.setStyleSheet("font-size: 14px;")
            analysis_layout.addWidget(self.avg_label, 1, 0)
            
            self.max_label = QLabel("Maksimum Tutar: 0 ₺")
            self.max_label.setStyleSheet("font-size: 14px;")
            analysis_layout.addWidget(self.max_label, 1, 1)
            
            analysis_group.setLayout(analysis_layout)
            layout.addWidget(analysis_group)
        
        # Butonlar (admin için düzenleme, üst düzey yönetici için sadece görüntüleme)
        button_layout = QHBoxLayout()
        
        # Test için manuel kayıt ekleme butonu
        test_add_btn = QPushButton("🧪 Test Kayıt Ekle")
        test_add_btn.clicked.connect(self.add_test_harcama_talep)
        button_layout.addWidget(test_add_btn)
        
        if self.role == 'admin':
            edit_btn = QPushButton("Düzenle")
            edit_btn.clicked.connect(self.edit_harcama_talep)
            button_layout.addWidget(edit_btn)
            
            delete_btn = QPushButton("Sil")
            delete_btn.clicked.connect(self.delete_harcama_talep)
            button_layout.addWidget(delete_btn)
            
            clear_all_btn = QPushButton("🗑️ Tümünü Temizle")
            clear_all_btn.setStyleSheet("background-color: #2563eb; color: white; font-weight: bold;")
            clear_all_btn.clicked.connect(self.clear_all_harcama_talep)
            button_layout.addWidget(clear_all_btn)
        
        export_btn = QPushButton("Excel'e Aktar")
        export_btn.clicked.connect(self.export_harcama_talep)
        button_layout.addWidget(export_btn)
        
        layout.addLayout(button_layout)
    
    def setup_masraf_tab(self):
        layout = QVBoxLayout(self.masraf_tab)
        
        # Tablo
        self.masraf_table = QTableWidget()
        self.masraf_table.setColumnCount(10)
        self.masraf_table.setHorizontalHeaderLabels([
            "ID", "Tarih", "Bölge Kodu", "Kaynak Tipi", "Stage",
            "Stage-Operasyon", "No.Su", "Kimden Alındığı", "Açıklama", "Tutar"
        ])
        self.masraf_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        layout.addWidget(self.masraf_table)
        
        # Analiz verileri
        if self.role in ['normal', 'ust_duzey_yonetici', 'admin']:
            analysis_group = QGroupBox("📊 Analiz Verileri")
            analysis_layout = QGridLayout()
            
            self.masraf_total_label = QLabel("Toplam Tutar: 0 ₺")
            self.masraf_total_label.setStyleSheet("font-size: 16px; font-weight: bold; color: #2563eb;")
            analysis_layout.addWidget(self.masraf_total_label, 0, 0)
            
            self.masraf_count_label = QLabel("Toplam Kayıt: 0")
            self.masraf_count_label.setStyleSheet("font-size: 14px;")
            analysis_layout.addWidget(self.masraf_count_label, 0, 1)
            
            self.masraf_avg_label = QLabel("Ortalama Tutar: 0 ₺")
            self.masraf_avg_label.setStyleSheet("font-size: 14px;")
            analysis_layout.addWidget(self.masraf_avg_label, 1, 0)
            
            self.masraf_max_label = QLabel("Maksimum Tutar: 0 ₺")
            self.masraf_max_label.setStyleSheet("font-size: 14px;")
            analysis_layout.addWidget(self.masraf_max_label, 1, 1)
            
            analysis_group.setLayout(analysis_layout)
            layout.addWidget(analysis_group)
        
        # Butonlar
        button_layout = QHBoxLayout()
        
        if self.role == 'admin':
            edit_btn = QPushButton("Düzenle")
            edit_btn.clicked.connect(self.edit_masraf)
            button_layout.addWidget(edit_btn)
            
            delete_btn = QPushButton("Sil")
            delete_btn.clicked.connect(self.delete_masraf)
            button_layout.addWidget(delete_btn)
            
            clear_all_btn = QPushButton("🗑️ Tümünü Temizle")
            clear_all_btn.setStyleSheet("background-color: #2563eb; color: white; font-weight: bold;")
            clear_all_btn.clicked.connect(self.clear_all_masraf)
            button_layout.addWidget(clear_all_btn)
        
        export_btn = QPushButton("Excel'e Aktar")
        export_btn.clicked.connect(self.export_masraf)
        button_layout.addWidget(export_btn)
        
        layout.addLayout(button_layout)
    
    def load_data(self):
        """Verileri yükle"""
        self.load_harcama_talep()
        self.load_masraf()
        self.update_filter_options()
    
    def update_filter_options(self):
        """Filtre combo box'larını güncelle"""
        # TEST İÇİN KALDIRILDI
        # Safha seçeneklerini yükle
        # try:
        #     url = f"http://127.0.0.1:5000/api/harcama_talep?user_id={self.user_id}"
        #     response = requests.get(url)
        #     if response.status_code == 200:
        #         data = response.json()
        #         if data.get('success'):
        #             expenses = data.get('data', [])
        #             safhalar = set()
        #             stage_kodlari = set()
        #             
        #             for expense in expenses:
        #                 if expense.get('safha'):
        #                     safhalar.add(expense.get('safha'))
        #                 if expense.get('stage_kodu'):
        #                     stage_kodlari.add(expense.get('stage_kodu'))
        #             
        #             # Safha combo box'ını güncelle
        #             if self.safha_combo:
        #                 self.safha_combo.clear()
        #                 self.safha_combo.addItem("Tümü", None)
        #                 for safha in sorted(safhalar):
        #                     self.safha_combo.addItem(safha, safha)
        #             
        #             # Stage combo box'ını güncelle
        #             if self.stage_combo:
        #                 self.stage_combo.clear()
        #                 self.stage_combo.addItem("Tümü", None)
        #                 for stage in sorted(stage_kodlari):
        #                     self.stage_combo.addItem(stage, stage)
        # except Exception as e:
        #     print(f"Filtre seçenekleri yüklenirken hata: {e}")
        pass
    
    def load_harcama_talep(self):
        """Harcama talep verilerini yükle"""
        try:
            # TEST İÇİN FİLTRELEME KALDIRILDI - Sadece user_id gönderiliyor
            params = {'user_id': self.user_id}
            
            # Filtre parametreleri kaldırıldı (test için)
            # bolge_kodu = self.bolge_combo.currentData() if self.bolge_combo else None
            # if bolge_kodu:
            #     params['bolge_kodu'] = bolge_kodu
            # 
            # safha = self.safha_combo.currentData() if self.safha_combo else None
            # if safha:
            #     params['safha'] = safha
            # 
            # stage_kodu = self.stage_combo.currentData() if self.stage_combo else None
            # if stage_kodu:
            #     params['stage_kodu'] = stage_kodu
            
            # URL oluştur
            url = "http://127.0.0.1:5000/api/harcama_talep"
            print(f"DEBUG - Harcama talep API çağrısı: {url}")
            print(f"DEBUG - Params: {params}")
            response = requests.get(url, params=params)
            
            print(f"DEBUG - Response status: {response.status_code}")
            if response.status_code == 200:
                data = response.json()
                print(f"DEBUG - Response data: {data}")
                if data.get('success'):
                    expenses = data.get('data', [])
                    print(f"DEBUG - Toplam {len(expenses)} harcama talep kaydı bulundu")
                    
                    self.harcama_table.setRowCount(len(expenses))
                    total = 0
                    for row_idx, expense in enumerate(expenses):
                        # ID kolonu (gizli)
                        self.harcama_table.setItem(row_idx, 0, QTableWidgetItem(str(expense.get('id', ''))))
                        self.harcama_table.setItem(row_idx, 1, QTableWidgetItem(str(expense.get('no', ''))))
                        self.harcama_table.setItem(row_idx, 2, QTableWidgetItem(str(expense.get('tarih', ''))))
                        self.harcama_table.setItem(row_idx, 3, QTableWidgetItem(str(expense.get('bolge_kodu', ''))))
                        self.harcama_table.setItem(row_idx, 4, QTableWidgetItem(str(expense.get('kaynak_tipi_kodu', ''))))
                        self.harcama_table.setItem(row_idx, 5, QTableWidgetItem(str(expense.get('stage_kodu', ''))))
                        self.harcama_table.setItem(row_idx, 6, QTableWidgetItem(str(expense.get('stage_operasyon_kodu', ''))))
                        self.harcama_table.setItem(row_idx, 7, QTableWidgetItem(str(expense.get('safha', ''))))
                        self.harcama_table.setItem(row_idx, 8, QTableWidgetItem(str(expense.get('harcama_kalemi', ''))))
                        self.harcama_table.setItem(row_idx, 9, QTableWidgetItem(str(expense.get('birim', ''))))
                        self.harcama_table.setItem(row_idx, 10, QTableWidgetItem(str(expense.get('miktar', ''))))
                        self.harcama_table.setItem(row_idx, 11, QTableWidgetItem(str(expense.get('birim_ucret', ''))))
                        self.harcama_table.setItem(row_idx, 12, QTableWidgetItem(str(expense.get('toplam', ''))))
                        self.harcama_table.setItem(row_idx, 13, QTableWidgetItem(str(expense.get('aciklama', ''))))
                        
                        # Toplam hesapla
                        toplam = expense.get('toplam', 0)
                        if toplam:
                            try:
                                total += float(toplam)
                            except:
                                pass
                    
                    # Analiz verilerini güncelle
                    if hasattr(self, 'total_label'):
                        self.total_label.setText(f"Toplam Tutar: {total:,.2f} ₺")
                        self.count_label.setText(f"Toplam Kayıt: {len(expenses)}")
                        if len(expenses) > 0:
                            avg = total / len(expenses)
                            self.avg_label.setText(f"Ortalama Tutar: {avg:,.2f} ₺")
                            
                            # Maksimum tutarı bul
                            max_val = 0
                            for expense in expenses:
                                toplam = expense.get('toplam', 0)
                                if toplam:
                                    try:
                                        max_val = max(max_val, float(toplam))
                                    except:
                                        pass
                            self.max_label.setText(f"Maksimum Tutar: {max_val:,.2f} ₺")
                        else:
                            self.avg_label.setText("Ortalama Tutar: 0 ₺")
                            self.max_label.setText("Maksimum Tutar: 0 ₺")
                else:
                    # API başarısız oldu
                    error_msg = data.get('message', 'Bilinmeyen hata')
                    print(f"API hatası: {error_msg}")
                    self.harcama_table.setRowCount(0)
            else:
                # HTTP hatası
                print(f"HTTP hatası: {response.status_code}")
                try:
                    error_data = response.json()
                    error_msg = error_data.get('message', f'HTTP {response.status_code} hatası')
                except:
                    error_msg = f'HTTP {response.status_code} hatası'
                print(f"API yanıt hatası: {error_msg}")
                self.harcama_table.setRowCount(0)
        except Exception as e:
            error_msg = f"Veriler yüklenirken hata oluştu: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", error_msg)
    
    def load_masraf(self):
        """Masraf verilerini yükle"""
        try:
            # TEST İÇİN FİLTRELEME KALDIRILDI - Sadece user_id gönderiliyor
            params = {'user_id': self.user_id}
            
            # Filtre parametreleri kaldırıldı (test için)
            # bolge_kodu = self.bolge_combo.currentData() if self.bolge_combo else None
            # if bolge_kodu:
            #     params['bolge_kodu'] = bolge_kodu
            # 
            # stage_kodu = self.stage_combo.currentData() if self.stage_combo else None
            # if stage_kodu:
            #     params['stage_kodu'] = stage_kodu
            
            # URL oluştur
            url = "http://127.0.0.1:5000/api/get_expenses"
            response = requests.get(url, params=params)
            
            if response.status_code == 200:
                data = response.json()
                if data.get('success'):
                    expenses = data.get('data', [])
                    
                    # API'de zaten filtrelenmiş veriler geldiği için direkt kullan
                    filtered_expenses = expenses
                    
                    self.masraf_table.setRowCount(len(filtered_expenses))
                    total = 0
                    for row_idx, expense in enumerate(filtered_expenses):
                        self.masraf_table.setItem(row_idx, 0, QTableWidgetItem(str(expense.get('id', ''))))
                        self.masraf_table.setItem(row_idx, 1, QTableWidgetItem(str(expense.get('tarih', ''))))
                        self.masraf_table.setItem(row_idx, 2, QTableWidgetItem(str(expense.get('bolge_kodu', ''))))
                        self.masraf_table.setItem(row_idx, 3, QTableWidgetItem(str(expense.get('kaynak_tipi', ''))))
                        self.masraf_table.setItem(row_idx, 4, QTableWidgetItem(str(expense.get('stage', ''))))
                        self.masraf_table.setItem(row_idx, 5, QTableWidgetItem(str(expense.get('stage_operasyon', ''))))
                        self.masraf_table.setItem(row_idx, 6, QTableWidgetItem(str(expense.get('no_su', ''))))
                        self.masraf_table.setItem(row_idx, 7, QTableWidgetItem(str(expense.get('kimden_alindigi', ''))))
                        self.masraf_table.setItem(row_idx, 8, QTableWidgetItem(str(expense.get('aciklama', ''))))
                        self.masraf_table.setItem(row_idx, 9, QTableWidgetItem(str(expense.get('tutar', ''))))
                        
                        # Toplam hesapla
                        tutar = expense.get('tutar', 0)
                        if tutar:
                            try:
                                total += float(tutar)
                            except:
                                pass
                    
                    # Analiz verilerini güncelle
                    if hasattr(self, 'masraf_total_label'):
                        self.masraf_total_label.setText(f"Toplam Tutar: {total:,.2f} ₺")
                        self.masraf_count_label.setText(f"Toplam Kayıt: {len(filtered_expenses)}")
                        if len(filtered_expenses) > 0:
                            avg = total / len(filtered_expenses)
                            self.masraf_avg_label.setText(f"Ortalama Tutar: {avg:,.2f} ₺")
                            
                            # Maksimum tutarı bul
                            max_val = 0
                            for expense in filtered_expenses:
                                tutar = expense.get('tutar', 0)
                                if tutar:
                                    try:
                                        max_val = max(max_val, float(tutar))
                                    except:
                                        pass
                            self.masraf_max_label.setText(f"Maksimum Tutar: {max_val:,.2f} ₺")
                        else:
                            self.masraf_avg_label.setText("Ortalama Tutar: 0 ₺")
                            self.masraf_max_label.setText("Maksimum Tutar: 0 ₺")
                else:
                    # API başarısız oldu
                    error_msg = data.get('message', 'Bilinmeyen hata')
                    print(f"API hatası: {error_msg}")
                    self.masraf_table.setRowCount(0)
            else:
                # HTTP hatası
                print(f"HTTP hatası: {response.status_code}")
                try:
                    error_data = response.json()
                    error_msg = error_data.get('message', f'HTTP {response.status_code} hatası')
                except:
                    error_msg = f'HTTP {response.status_code} hatası'
                print(f"API yanıt hatası: {error_msg}")
                self.masraf_table.setRowCount(0)
        except Exception as e:
            error_msg = f"Masraf verileri yüklenirken hata oluştu: {str(e)}"
            print(error_msg)
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Hata", error_msg)
    
    def apply_filters(self):
        """Filtreleri uygula"""
        self.load_harcama_talep()
        self.load_masraf()
    
    def add_test_harcama_talep(self):
        """Test için manuel harcama talep kaydı ekle"""
        try:
            from datetime import datetime
            import requests
            
            # Test verisi
            test_data = {
                'user_id': self.user_id,
                'tarih': datetime.now().strftime('%Y-%m-%d'),
                'bolge_kodu': '10',
                'kaynak_tipi_kodu': '01',
                'stage_kodu': '01',
                'stage_operasyon_kodu': '0101',
                'safha': 'Fidelik',
                'harcama_kalemi': 'Fide Yastığı Hazırlama',
                'birim': 'YEVMİYE',
                'miktar': 10.0,
                'birim_ucret': 1250.0,
                'toplam': 12500.0,
                'aciklama': 'Test kaydı - Manuel ekleme',
                'is_manuel': 1
            }
            
            print(f"DEBUG - Test kayıt ekleniyor: {test_data}")
            response = requests.post('http://127.0.0.1:5000/api/harcama_talep', json=test_data, timeout=5)
            
            print(f"DEBUG - Test kayıt yanıtı: {response.status_code} - {response.text}")
            
            if response.status_code == 201:
                QMessageBox.information(self, "Başarılı", "Test kaydı başarıyla eklendi!")
                self.load_harcama_talep()  # Verileri yenile
            else:
                QMessageBox.warning(self, "Hata", f"Kayıt eklenemedi: {response.text}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Test kaydı eklenirken hata: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def edit_harcama_talep(self):
        """Harcama talep düzenle"""
        selected_row = self.harcama_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen düzenlemek istediğiniz satırı seçin.")
            return
        
        # Düzenleme dialog'u oluştur
        dialog = QDialog(self)
        dialog.setWindowTitle("Harcama Talep Düzenle")
        dialog.setMinimumWidth(500)
        
        layout = QFormLayout(dialog)
        
        # ID'yi al (gizli kolon)
        harcama_talep_id = int(self.harcama_table.item(selected_row, 0).text())
        
        # Mevcut değerleri al (kolonlar 1 kaydı)
        no = self.harcama_table.item(selected_row, 1).text()
        tarih = self.harcama_table.item(selected_row, 2).text()
        bolge_kodu = self.harcama_table.item(selected_row, 3).text()
        kaynak_tipi_kodu = self.harcama_table.item(selected_row, 4).text()
        stage_kodu = self.harcama_table.item(selected_row, 5).text()
        stage_operasyon_kodu = self.harcama_table.item(selected_row, 6).text()
        safha = self.harcama_table.item(selected_row, 7).text()
        harcama_kalemi = self.harcama_table.item(selected_row, 8).text()
        birim = self.harcama_table.item(selected_row, 9).text()
        miktar = self.harcama_table.item(selected_row, 10).text()
        birim_ucret = self.harcama_table.item(selected_row, 11).text()
        toplam = self.harcama_table.item(selected_row, 12).text()
        aciklama = self.harcama_table.item(selected_row, 13).text()
        
        # Form alanları
        tarih_edit = QDateEdit()
        try:
            tarih_edit.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        except:
            tarih_edit.setDate(QDate.currentDate())
        
        bolge_edit = QLineEdit(bolge_kodu)
        kaynak_tipi_edit = QLineEdit(kaynak_tipi_kodu)
        stage_edit = QLineEdit(stage_kodu)
        stage_operasyon_edit = QLineEdit(stage_operasyon_kodu)
        safha_edit = QLineEdit(safha)
        harcama_kalemi_edit = QLineEdit(harcama_kalemi)
        birim_edit = QLineEdit(birim)
        miktar_edit = QDoubleSpinBox()
        miktar_edit.setMaximum(999999999)
        miktar_edit.setValue(float(miktar) if miktar else 0)
        birim_ucret_edit = QDoubleSpinBox()
        birim_ucret_edit.setMaximum(999999999)
        birim_ucret_edit.setValue(float(birim_ucret) if birim_ucret else 0)
        toplam_edit = QDoubleSpinBox()
        toplam_edit.setMaximum(999999999)
        toplam_edit.setValue(float(toplam) if toplam else 0)
        aciklama_edit = QLineEdit(aciklama)
        
        layout.addRow("Tarih:", tarih_edit)
        layout.addRow("Bölge Kodu:", bolge_edit)
        layout.addRow("Kaynak Tipi Kodu:", kaynak_tipi_edit)
        layout.addRow("Stage Kodu:", stage_edit)
        layout.addRow("Stage-Operasyon Kodu:", stage_operasyon_edit)
        layout.addRow("Safha:", safha_edit)
        layout.addRow("Harcama Kalemi:", harcama_kalemi_edit)
        layout.addRow("Birim:", birim_edit)
        layout.addRow("Miktar:", miktar_edit)
        layout.addRow("Birim Ücret:", birim_ucret_edit)
        layout.addRow("Toplam:", toplam_edit)
        layout.addRow("Açıklama:", aciklama_edit)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec_() == QDialog.Accepted:
            # API'ye güncelleme isteği gönder
            try:
                update_data = {
                    'user_id': self.user_id,
                    'tarih': tarih_edit.date().toString('yyyy-MM-dd'),
                    'bolge_kodu': bolge_edit.text(),
                    'kaynak_tipi_kodu': kaynak_tipi_edit.text(),
                    'stage_kodu': stage_edit.text(),
                    'stage_operasyon_kodu': stage_operasyon_edit.text(),
                    'safha': safha_edit.text(),
                    'harcama_kalemi': harcama_kalemi_edit.text(),
                    'birim': birim_edit.text(),
                    'miktar': miktar_edit.value(),
                    'birim_ucret': birim_ucret_edit.value(),
                    'toplam': toplam_edit.value(),
                    'aciklama': aciklama_edit.text()
                }
                
                url = f"http://127.0.0.1:5000/api/harcama_talep/{harcama_talep_id}"
                response = requests.put(url, json=update_data, timeout=10)
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Başarılı", "Harcama talep başarıyla güncellendi!")
                    self.load_harcama_talep()  # Verileri yenile
                else:
                    error_msg = response.json().get('message', 'Bilinmeyen hata')
                    QMessageBox.warning(self, "Hata", f"Güncelleme başarısız: {error_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Güncelleme sırasında hata oluştu: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def delete_harcama_talep(self):
        """Harcama talep sil"""
        selected_row = self.harcama_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek istediğiniz satırı seçin.")
            return
        
        reply = QMessageBox.question(self, "Onay", "Bu kaydı silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                # ID'yi al (gizli kolon)
                harcama_talep_id = int(self.harcama_table.item(selected_row, 0).text())
                
                # API'ye silme isteği gönder
                url = f"http://127.0.0.1:5000/api/harcama_talep/{harcama_talep_id}"
                response = requests.delete(url, timeout=10)
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Başarılı", "Harcama talep başarıyla silindi!")
                    self.load_harcama_talep()  # Verileri yenile
                else:
                    error_msg = response.json().get('message', 'Bilinmeyen hata')
                    QMessageBox.warning(self, "Hata", f"Silme başarısız: {error_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Silme sırasında hata oluştu: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def clear_all_harcama_talep(self):
        """Tüm harcama talep kayıtlarını temizle (admin için)"""
        reply = QMessageBox.question(
            self, 
            "Onay", 
            "TÜM harcama talep kayıtlarını silmek istediğinizden emin misiniz?\n\nBu işlem geri alınamaz!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                url = "http://127.0.0.1:5000/api/clear_harcama_talep"
                response = requests.delete(url, timeout=10)
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Başarılı", "Tüm harcama talep kayıtları başarıyla silindi!")
                    self.load_harcama_talep()  # Verileri yenile
                else:
                    error_msg = response.json().get('message', 'Bilinmeyen hata')
                    QMessageBox.warning(self, "Hata", f"Temizleme başarısız: {error_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Temizleme sırasında hata oluştu: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def clear_all_masraf(self):
        """Tüm masraf kayıtlarını temizle (admin için)"""
        reply = QMessageBox.question(
            self, 
            "Onay", 
            "TÜM masraf kayıtlarını silmek istediğinizden emin misiniz?\n\nBu işlem geri alınamaz!",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            try:
                url = "http://127.0.0.1:5000/api/clear_all_expenses"
                response = requests.delete(url, timeout=10)
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Başarılı", "Tüm masraf kayıtları başarıyla silindi!")
                    self.load_masraf()  # Verileri yenile
                else:
                    error_msg = response.json().get('message', 'Bilinmeyen hata')
                    QMessageBox.warning(self, "Hata", f"Temizleme başarısız: {error_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Temizleme sırasında hata oluştu: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def export_harcama_talep(self):
        """Excel'e aktar"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Excel Dosyası Kaydet", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Tablodan veri al (ID kolonunu atla - ilk kolon)
            data = []
            for row in range(self.harcama_table.rowCount()):
                row_data = []
                for col in range(1, self.harcama_table.columnCount()):  # 1'den başla (ID'yi atla)
                    item = self.harcama_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # DataFrame oluştur
            df = pd.DataFrame(data, columns=[
                "No", "Tarih", "BÖLGE KODU", "KAYNAK TİPİ KODU", "STAGE KODU",
                "STAGE-OPERASYON KODU", "Safha", "Harcama Kalemi", "Birim",
                "Miktar", "Birim ücret", "Toplam", "Açıklama"
            ])
            
            # Excel'e yaz
            wb = Workbook()
            ws = wb.active
            
            # Başlık
            ws['A1'] = "Harcama Talep Raporu"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:M1')
            
            # Başlık satırı
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Veri satırları
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    # Birim ücret ve Toplam kolonlarına ₺ ekle
                    if col_idx == 11 or col_idx == 12:  # Birim ücret ve Toplam
                        if value and str(value).replace('.', '').replace('-', '').isdigit():
                            cell.value = f"{float(value):,.2f} ₺"
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['M'].width = 30
            
            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", f"Excel dosyası kaydedildi: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel'e aktarma hatası: {str(e)}")
    
    def edit_masraf(self):
        """Masraf düzenle"""
        selected_row = self.masraf_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen düzenlemek istediğiniz satırı seçin.")
            return
        
        # Düzenleme dialog'u oluştur
        dialog = QDialog(self)
        dialog.setWindowTitle("Masraf Düzenle")
        dialog.setMinimumWidth(500)
        
        layout = QFormLayout(dialog)
        
        # ID'yi al
        expense_id = int(self.masraf_table.item(selected_row, 0).text())
        
        # Mevcut değerleri al
        tarih = self.masraf_table.item(selected_row, 1).text()
        bolge_kodu = self.masraf_table.item(selected_row, 2).text()
        kaynak_tipi = self.masraf_table.item(selected_row, 3).text()
        stage = self.masraf_table.item(selected_row, 4).text()
        stage_operasyon = self.masraf_table.item(selected_row, 5).text()
        no_su = self.masraf_table.item(selected_row, 6).text()
        kimden_alindigi = self.masraf_table.item(selected_row, 7).text()
        aciklama = self.masraf_table.item(selected_row, 8).text()
        tutar = self.masraf_table.item(selected_row, 9).text()
        
        # Form alanları
        tarih_edit = QDateEdit()
        try:
            tarih_edit.setDate(QDate.fromString(tarih, "yyyy-MM-dd"))
        except:
            tarih_edit.setDate(QDate.currentDate())
        
        bolge_edit = QLineEdit(bolge_kodu)
        kaynak_tipi_edit = QLineEdit(kaynak_tipi)
        stage_edit = QLineEdit(stage)
        stage_operasyon_edit = QLineEdit(stage_operasyon)
        no_su_edit = QLineEdit(no_su)
        kimden_alindigi_edit = QLineEdit(kimden_alindigi)
        aciklama_edit = QLineEdit(aciklama)
        tutar_edit = QDoubleSpinBox()
        tutar_edit.setMaximum(999999999)
        tutar_edit.setValue(float(tutar) if tutar else 0)
        
        layout.addRow("Tarih:", tarih_edit)
        layout.addRow("Bölge Kodu:", bolge_edit)
        layout.addRow("Kaynak Tipi:", kaynak_tipi_edit)
        layout.addRow("Stage:", stage_edit)
        layout.addRow("Stage-Operasyon:", stage_operasyon_edit)
        layout.addRow("No.Su:", no_su_edit)
        layout.addRow("Kimden Alındığı:", kimden_alindigi_edit)
        layout.addRow("Açıklama:", aciklama_edit)
        layout.addRow("Tutar:", tutar_edit)
        
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        layout.addRow(buttons)
        
        if dialog.exec_() == QDialog.Accepted:
            # API'ye güncelleme isteği gönder
            try:
                update_data = {
                    'tarih': tarih_edit.date().toString('yyyy-MM-dd'),
                    'bolge_kodu': bolge_edit.text(),
                    'kaynak_tipi': kaynak_tipi_edit.text(),
                    'stage': stage_edit.text(),
                    'stage_operasyon': stage_operasyon_edit.text(),
                    'no_su': no_su_edit.text(),
                    'kimden_alindigi': kimden_alindigi_edit.text(),
                    'aciklama': aciklama_edit.text(),
                    'tutar': tutar_edit.value()
                }
                
                url = f"http://127.0.0.1:5000/api/update_expense/{expense_id}"
                response = requests.put(url, json=update_data, timeout=10)
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Başarılı", "Masraf başarıyla güncellendi!")
                    self.load_masraf()  # Verileri yenile
                else:
                    error_msg = response.json().get('message', 'Bilinmeyen hata')
                    QMessageBox.warning(self, "Hata", f"Güncelleme başarısız: {error_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Güncelleme sırasında hata oluştu: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def delete_masraf(self):
        """Masraf sil"""
        selected_row = self.masraf_table.currentRow()
        if selected_row < 0:
            QMessageBox.warning(self, "Uyarı", "Lütfen silmek istediğiniz satırı seçin.")
            return
        
        reply = QMessageBox.question(self, "Onay", "Bu kaydı silmek istediğinizden emin misiniz?",
                                     QMessageBox.Yes | QMessageBox.No)
        
        if reply == QMessageBox.Yes:
            try:
                # ID'yi al
                expense_id = int(self.masraf_table.item(selected_row, 0).text())
                
                # API'ye silme isteği gönder
                url = f"http://127.0.0.1:5000/api/delete_expense/{expense_id}"
                response = requests.delete(url, timeout=10)
                
                if response.status_code == 200:
                    QMessageBox.information(self, "Başarılı", "Masraf başarıyla silindi!")
                    self.load_masraf()  # Verileri yenile
                else:
                    error_msg = response.json().get('message', 'Bilinmeyen hata')
                    QMessageBox.warning(self, "Hata", f"Silme başarısız: {error_msg}")
            except Exception as e:
                QMessageBox.critical(self, "Hata", f"Silme sırasında hata oluştu: {str(e)}")
                import traceback
                traceback.print_exc()
    
    def export_masraf(self):
        """Excel'e aktar"""
        try:
            file_path, _ = QFileDialog.getSaveFileName(
                self, "Excel Dosyası Kaydet", "", "Excel Files (*.xlsx)"
            )
            
            if not file_path:
                return
            
            # Tablodan veri al
            data = []
            for row in range(self.masraf_table.rowCount()):
                row_data = []
                for col in range(self.masraf_table.columnCount()):
                    item = self.masraf_table.item(row, col)
                    row_data.append(item.text() if item else "")
                data.append(row_data)
            
            # DataFrame oluştur
            df = pd.DataFrame(data, columns=[
                "ID", "Tarih", "Bölge Kodu", "Kaynak Tipi", "Stage",
                "Stage-Operasyon", "No.Su", "Kimden Alındığı", "Açıklama", "Tutar"
            ])
            
            # Excel'e yaz
            wb = Workbook()
            ws = wb.active
            
            # Başlık
            ws['A1'] = "Masraf Raporu"
            ws['A1'].font = Font(bold=True, size=16)
            ws.merge_cells('A1:J1')
            
            # Başlık satırı
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True, color="000000")
                cell.fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Veri satırları
            for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    # Tutar kolonuna ₺ ekle
                    if col_idx == 10:  # Tutar
                        if value and str(value).replace('.', '').replace('-', '').isdigit():
                            cell.value = f"{float(value):,.2f} ₺"
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['I'].width = 30
            
            wb.save(file_path)
            QMessageBox.information(self, "Başarılı", f"Excel dosyası kaydedildi: {file_path}")
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Excel'e aktarma hatası: {str(e)}")
