import pandas as pd
import numpy as np
from PyQt5.QtWidgets import (QApplication, QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QPushButton, QFileDialog, QMessageBox, QTableWidget,
                             QTableWidgetItem, QHeaderView, QProgressBar, QDialog,
                             QFormLayout, QComboBox, QLineEdit, QDateEdit, QDoubleSpinBox,
                             QDialogButtonBox)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate
import sys
from datetime import datetime
import openpyxl
import re
from difflib import SequenceMatcher
import traceback
import sys
import os
import logging
import sys
import os
import logging
from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QLabel,
                             QPushButton, QProgressBar, QTableWidget,
                             QTableWidgetItem, QFileDialog, QMessageBox)
from PyQt5.QtCore import Qt
import requests
from api_auth_context import merge_auth_headers
from config import get_api_root

def ensure_class_availability():
    """PyInstaller paketlemesinde sınıf referanslarının kaybolmaması için"""
    try:
        # ExcelProcessorThread sınıfının globals'da mevcut olduğundan emin ol
        if 'ExcelProcessorThread' in globals():
            logging.info("ExcelProcessorThread sınıfı globals'da mevcut")
            return True
        else:
            logging.error("ExcelProcessorThread sınıfı globals'da bulunamadı")
            return False
    except Exception as e:
        logging.error(f"Sınıf kontrol hatası: {e}")
        return False


def get_resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
        logging.info(f"PyInstaller bundle path: {base_path}")
    except Exception:
        base_path = os.path.abspath(".")
        logging.info(f"Development path: {base_path}")

    full_path = os.path.join(base_path, relative_path)
    logging.info(f"Resource path: {full_path}")
    return full_path


# Fuzzy matching için alternatif yaklaşım
def simple_fuzzy_ratio(s1, s2):
    """Basit fuzzy matching oranı hesaplar"""
    if not s1 or not s2:
        return 0

    s1, s2 = s1.lower(), s2.lower()
    if s1 == s2:
        return 100

    # Sequence matcher kullan
    matcher = SequenceMatcher(None, s1, s2)
    return int(matcher.ratio() * 100)


def simple_partial_ratio(s1, s2):
    """Basit kısmi eşleşme oranı"""
    if not s1 or not s2:
        return 0

    s1, s2 = s1.lower(), s2.lower()
    shorter = s1 if len(s1) <= len(s2) else s2
    longer = s2 if len(s1) <= len(s2) else s1

    if shorter in longer:
        return 100

    best_ratio = 0
    for i in range(len(longer) - len(shorter) + 1):
        substring = longer[i:i + len(shorter)]
        ratio = simple_fuzzy_ratio(shorter, substring)
        best_ratio = max(best_ratio, ratio)

    return best_ratio


class ExcelProcessorThread(QThread):
    """
    Excel dosyasını arka planda işlemek için kullanılan iş parçacığı sınıfı.
    İlerleme güncellemeleri, işleme tamamlandığında işlenmiş DataFrame ve hata mesajları için sinyaller sağlar.
    """
    progress_updated = pyqtSignal(int)
    #processing_finished = pyqtSignal(pd.DataFrame)
    processing_finished = pyqtSignal(object)
    error_occurred = pyqtSignal(str)

    def __init__(self, input_file, api_client):
        super().__init__()
        print("DEBUG - ExcelProcessorThread başlatılıyor")

        self.input_file = input_file
        self.api_client = api_client
        self.mappings = {}
        self._is_running = True  # Çalışma durumu kontrolü
        
        print("DEBUG - ExcelProcessorThread başlatıldı")

    def run(self):
        try:
            print("DEBUG - Thread run metodu başlatılıyor...")
            self.progress_updated.emit(10)

            # Thread'in durumunu kontrol et
            if not self._is_running or self.isInterruptionRequested():
                return

            if not self.load_mappings():
                self.error_occurred.emit("Veritabanından veriler yüklenemedi!")
                return

            print("DEBUG - Excel dosyası okunuyor...")
            self.progress_updated.emit(20)

            # Dosya formatına göre engine seç (try bloğundan önce tanımla)
            import os
            file_ext = os.path.splitext(self.input_file)[1].lower()
            
            # Excel okuma için engine belirle
            if file_ext == '.xls':
                # Eski .xls formatı için xlrd kullan (varsa)
                try:
                    import xlrd
                    engine = 'xlrd'
                except ImportError:
                    # xlrd yoksa hata ver
                    self.error_occurred.emit(
                        "xlrd modülü bulunamadı. .xls dosyaları için gerekli.\n"
                        "Yüklemek için: pip install xlrd>=2.0.1")
                    return
            else:
                # .xlsx ve diğer formatlar için openpyxl kullan
                engine = 'openpyxl'
            
            print(f"DEBUG - Excel dosya formatı: {file_ext}, Engine: {engine}")

            try:
                # Excel'i başlıkları olmadan oku
                df_temp = pd.read_excel(self.input_file, header=None, dtype=str, engine=engine)

                # Temel başlıklar (bunlar mutlaka olmalı)
                required_headers = {"Saha", "Alt Saha", "Safha", "Operasyon", "Üreti Tarihi", "Isgücü Tipi",
                                    "Ürün Miktar"}

                # Opsiyonel başlık (varsa dahil edilir)
                optional_headers = {"Bilgi"}

                # Tüm olası başlıklar
                all_possible_headers = required_headers.union(optional_headers)

                header_row_index = -1
                found_headers = set()
                best_match_score = 0
                best_match_row = -1

                # İlk 20 satırı kontrol et (daha geniş arama)
                for i, row in df_temp.head(20).iterrows():
                    # Satırdaki tüm değerleri temizle ve liste'ye çevir (sıra önemli)
                    row_values_list = []
                    row_values_set = set()
                    for val in row.dropna():
                        if pd.notna(val) and str(val).strip():
                            # Başlık değerlerini normalize et
                            normalized_val = str(val).strip()
                            row_values_list.append(normalized_val)
                            row_values_set.add(normalized_val)

                    print(f"DEBUG - Satır {i} başlıkları: {row_values_list}")

                    # Bu satırda kaç tane gerekli başlık var?
                    matching_headers = set()
                    for header in required_headers:
                        # Önce exact match kontrolü
                        if header in row_values_set:
                            matching_headers.add(header)
                        else:
                            # Alternatif yazımları da kontrol et (tüm satır değerlerini kontrol et)
                            for row_val in row_values_list:
                                if self.is_header_match(header, row_val):
                                    matching_headers.add(header)
                                    print(f"DEBUG - Satır {i}: '{header}' eşleşti '{row_val}' ile")
                                    break

                    print(f"DEBUG - Satır {i} eşleşen başlıklar ({len(matching_headers)}/{len(required_headers)}): {matching_headers}")

                    # Eşleşme skorunu hesapla
                    match_score = len(matching_headers)
                    
                    # En iyi eşleşmeyi kaydet
                    if match_score > best_match_score:
                        best_match_score = match_score
                        best_match_row = i
                        found_headers = matching_headers

                    # Eğer gerekli başlıkların tamamı veya büyük çoğunluğu varsa (%70 eşleşme yeterli)
                    if match_score >= len(required_headers) * 0.70:
                        header_row_index = i
                        print(f"DEBUG - Başlık satırı bulundu: Satır {i} ({match_score}/{len(required_headers)} başlık)")
                        break

                # Eğer %70 eşleşme bulunamadıysa, en iyi eşleşmeyi kullan
                if header_row_index == -1 and best_match_score >= 4:  # En az 4 temel başlık
                    header_row_index = best_match_row
                    print(f"DEBUG - En iyi eşleşme kullanılıyor: Satır {best_match_row} ({best_match_score}/{len(required_headers)} başlık)")
                elif header_row_index == -1:
                    # Son deneme: Daha esnek arama
                    print("DEBUG - Tam eşleşme bulunamadı, çok esnek eşleşme deneniyor...")
                    for i, row in df_temp.head(20).iterrows():
                        row_values = [str(val).strip() for val in row.dropna() if pd.notna(val) and str(val).strip()]

                        # En az 3 temel başlık varsa kabul et
                        matching_count = 0
                        for header in required_headers:
                            for row_val in row_values:
                                if self.is_header_match(header, row_val):
                                    matching_count += 1
                                    break

                        if matching_count >= 3:  # En az 3 temel başlık
                            header_row_index = i
                            print(f"DEBUG - Çok esnek eşleşme ile başlık satırı bulundu: Satır {i} ({matching_count} başlık)")
                            break

                if header_row_index == -1:
                    self.error_occurred.emit(
                        "Başlık satırı bulunamadı. Lütfen Excel dosyasında şu başlıkların olduğundan emin olun:\n" +
                        ", ".join(required_headers))
                    return

                # Dosyayı doğru başlık satırıyla tekrar oku (aynı engine'i kullan)
                df = pd.read_excel(self.input_file, header=header_row_index, dtype=str, engine=engine)

                # Sütun isimlerindeki olası boşlukları temizle
                df.columns = df.columns.str.strip()

                print(f"DEBUG - Okunan tüm sütunlar: {df.columns.tolist()}")

                # Gerekli sütunları bul ve eşleştir
                column_mapping = {}
                final_columns_to_keep = []

                for required_header in required_headers:
                    found_column = None
                    for col in df.columns:
                        if self.is_header_match(required_header, col):
                            found_column = col
                            break

                    if found_column:
                        column_mapping[found_column] = required_header
                        final_columns_to_keep.append(found_column)
                        print(f"DEBUG - Eşleşme: '{found_column}' -> '{required_header}'")
                    else:
                        print(f"WARNING - Gerekli başlık bulunamadı: {required_header}")

                # Opsiyonel başlıkları da kontrol et
                for optional_header in optional_headers:
                    found_column = None
                    for col in df.columns:
                        if self.is_header_match(optional_header, col):
                            found_column = col
                            break

                    if found_column:
                        column_mapping[found_column] = optional_header
                        final_columns_to_keep.append(found_column)
                        print(f"DEBUG - Opsiyonel eşleşme: '{found_column}' -> '{optional_header}'")

                # Eğer çok az sütun eşleştiyse hata ver
                if len(final_columns_to_keep) < 5:
                    missing_headers = required_headers - set(column_mapping.values())
                    self.error_occurred.emit(
                        f"Yeterli başlık bulunamadı. Eksik başlıklar: {', '.join(missing_headers)}")
                    return

                # Sadece eşleştirilen sütunları al
                df = df[final_columns_to_keep]

                # Sütun isimlerini standart isimlere çevir
                df = df.rename(columns=column_mapping)

                print(f"DEBUG - İşlenecek standart sütunlar: {df.columns.tolist()}")

            except Exception as e:
                import traceback
                error_msg = f"Excel dosyası işlenirken hata oluştu: {str(e)}\n{traceback.format_exc()}"
                self.error_occurred.emit(error_msg)
                return

            if df.empty:
                self.error_occurred.emit("Excel dosyası boş veya işlenecek veri bulunamadı!")
                return

            print("DEBUG - Veri işleme başlıyor...")
            self.progress_updated.emit(30)

            processed_df = self.process_data(df)
            if processed_df is None or processed_df.empty:
                self.error_occurred.emit("Veri işleme başarısız!")
                return

            print("DEBUG - Veri gruplama başlıyor...")
            self.progress_updated.emit(80)

            grouped_df = self.group_data(processed_df)
            if grouped_df is None or grouped_df.empty:
                self.error_occurred.emit("Veri gruplama başarısız!")
                return

            print(f"DEBUG - İşlem tamamlandı: {len(grouped_df)} satır")
            self.progress_updated.emit(100)
            
            # Eğer thread hala çalışıyorsa sinyal gönder
            if self._is_running and not self.isInterruptionRequested():
                print("DEBUG - processing_finished sinyali gönderiliyor")
                print(f"DEBUG - Gönderilecek DataFrame: {type(grouped_df)}, boyut: {grouped_df.shape if grouped_df is not None else 'None'}")
                print(f"DEBUG - Thread durumu: isRunning={self.isRunning()}, isFinished={self.isFinished()}")
                
                # Sinyali gönder
                try:
                    self.processing_finished.emit(grouped_df)
                    print("DEBUG - processing_finished sinyali başarıyla gönderildi")
                except Exception as e:
                    print(f"DEBUG - Sinyal gönderme hatası: {e}")
                    import traceback
                    print(f"DEBUG - Sinyal gönderme traceback: {traceback.format_exc()}")
                
                # Sinyalin işlenmesi için kısa bir süre bekle
                import time
                time.sleep(0.2)
                print("DEBUG - Sinyal gönderme sonrası bekleme tamamlandı")
                
            else:
                print("DEBUG - Thread durdurulmuş, sinyal gönderilmiyor")
            
            print("DEBUG - Thread run metodu tamamlandı")
            
            # Thread'in tamamen bitmesini bekle
            print("DEBUG - Thread'in tamamen bitmesi bekleniyor...")
            self.wait(1000)  # 1 saniye bekle
            print("DEBUG - Thread tamamen bitti")

            # Thread'in durumunu kontrol et
            print(f"DEBUG - Thread final durumu: isRunning={self.isRunning()}, isFinished={self.isFinished()}")
            
            # Uygulamanın kapanmadığından emin ol
            try:
                app = QApplication.instance()
                if app:
                    print("DEBUG - QApplication instance hala mevcut")
                    if hasattr(app, 'main_window') and app.main_window:
                        print("DEBUG - Ana pencere hala mevcut")
                    else:
                        print("DEBUG - Ana pencere bulunamadı")
                else:
                    print("DEBUG - QApplication instance bulunamadı")
            except Exception as e:
                print(f"DEBUG - Uygulama kontrol hatası: {e}")

        except Exception as e:
            import traceback
            error_msg = f"Beklenmeyen hata: {str(e)}\n{traceback.format_exc()}"
            print(f"CRITICAL ERROR - {error_msg}")
            self.error_occurred.emit(error_msg)

    def stop(self):
        """Thread'i güvenli şekilde durdur"""
        self._is_running = False
        self.requestInterruption()
    def is_header_match(self, expected_header, actual_header):
        """Başlık eşleşmesini kontrol eder - esnek karşılaştırma"""
        try:
            if not expected_header or not actual_header:
                return False

            expected = str(expected_header).strip().upper()
            actual = str(actual_header).strip().upper()

            # Tam eşleşme
            if expected == actual:
                return True

            # Türkçe karakter normalizasyonu
            turkish_chars = {
                'Ç': 'C', 'Ğ': 'G', 'İ': 'I', 'Ö': 'O', 'Ş': 'S', 'Ü': 'U',
                'ç': 'C', 'ğ': 'G', 'ı': 'I', 'ö': 'O', 'ş': 'S', 'ü': 'U'
            }

            for tr_char, en_char in turkish_chars.items():
                expected = expected.replace(tr_char, en_char)
                actual = actual.replace(tr_char, en_char)

            # Normalize edilmiş tam eşleşme
            if expected == actual:
                return True

            # Özel eşleşmeler
            special_matches = {
                'SAHA': ['SAHA', 'FIELD'],
                'ALT SAHA': ['ALT SAHA', 'SUB FIELD', 'ALTSAHA', 'ALT_SAHA'],
                'SAFHA': ['SAFHA', 'PHASE', 'STAGE'],
                'OPERASYON': ['OPERASYON', 'OPERATION', 'ISLEM', 'IŞLEM'],
                'URETI TARIHI': ['URETI TARIHI', 'URETIM TARIHI', 'PRODUCTION DATE', 'TARIH', 'DATE'],
                'ISGUCU TIPI': ['ISGUCU TIPI', 'IŞGÜCÜ TIPI', 'WORKFORCE TYPE', 'WORKER TYPE', 'CALISAN TIPI',
                                'ÇALIŞAN TIPI'],
                'URUN MIKTAR': ['URUN MIKTAR', 'ÜRÜN MIKTAR', 'PRODUCT AMOUNT', 'MIKTAR', 'AMOUNT', 'QUANTITY'],
                'BILGI': ['BILGI', 'INFO', 'INFORMATION', 'NOTE', 'NOTES', 'ACIKLAMA', 'AÇIKLAMA']
            }

            if expected in special_matches:
                return actual in special_matches[expected]

            # Kısmi eşleşme (içeriyor mu?)
            if expected in actual or actual in expected:
                return True

            return False

        except Exception as e:
            print(f"WARNING - Header match kontrolünde hata: {str(e)}")
            return False

    def load_mappings(self):
        """Veritabanından gerekli mapping'leri yükler"""
        try:
            print("DEBUG - Mappings yükleniyor...")

            # API client kontrolü
            if not self.api_client:
                print("ERROR - API client bulunamadı!")
                return False

            # Tüm verileri güvenli şekilde al
            try:
                all_data = self.api_client.get_all_data()
            except Exception as e:
                print(f"ERROR - API'den veri alınamadı: {str(e)}")
                return False

            if not all_data:
                print("ERROR - API'den boş veri döndü!")
                return False

            # Mappings'i güvenli şekilde yükle
            self.mappings = {
                'stage_kodlari': all_data.get('stages', {}),
                'operasyonlar': all_data.get('operasyonlar', {}),
                'stage_operasyonlar': all_data.get('stage_operasyonlar', {}),
                'kaynak_tipleri': all_data.get('kaynak_tipleri', {}),
                'bolge_kodlari': all_data.get('bolge_kodlari', {}),
                'birim_ucretler': all_data.get('birim_ucretler', {})
            }

            # Mappings kontrolü
            for key, value in self.mappings.items():
                if not isinstance(value, dict):
                    print(f"WARNING - {key} mapping'i dict değil: {type(value)}")
                    self.mappings[key] = {}

            print(f"DEBUG - Mappings yüklendi:")
            for key, value in self.mappings.items():
                print(f"  {key}: {len(value)} adet")
                if key == 'bolge_kodlari' and value:
                    print(f"    Bölge kodları örnekleri: {dict(list(value.items())[:5])}")

            return True

        except Exception as e:
            print(f"ERROR - Mapping yüklenirken hata: {str(e)}")
            return False

    def create_fuzzy_match_dictionary(self):
        """EN ÜST SEVİYE yaygın yazım hatalarını ve varyasyonlarını içeren genişletilmiş sözlük"""
        return {
            # Aşılama varyasyonları
            'ASILAMA': ['ASILAMA', 'ASIM', 'ASILAMA ISCILIK', 'ASILAMA ISCILIGI', 'ASILAMA IŞCILIK', 
                        'ASILAMA IŞÇILIK', 'ASILAMA ISCI', 'ASIM ISCILIK', 'ASIM ISCI', 'AŞILAMA', 
                        'AŞILAMA ISCILIK', 'AŞILAMA IŞÇILIK'],

            # Sulama varyasyonları
            'SULAMA': ['SULAMA', 'SULAMA ISCILIK', 'SULAMA ISCILIGI', 'SULAMA SISTEMI', 'SULAMA IŞCILIK',
                       'SULAMA IŞÇILIK', 'SULAMA ISCI', 'SULAMA SISTEM', 'SULAMA SISTEMI ISCILIK',
                       'SULAMA SISTEM ISCILIK', 'SULAMA KURULUMU', 'SULAMA KURULUM'],
            'SULAMA TAMIR': ['SULAMA TAMIR', 'SULAMA TAMIRAT', 'SULAMA BAKIM', 'SULAMA SISTEMI TAMIR',
                             'SULAMA SISTEMI TAMIRAT', 'SULAMA SISTEM TAMIR', 'SULAMA SISTEM TAMIRAT',
                             'SULAMA TAMIR ISCILIK', 'SULAMA TAMIRAT ISCILIK', 'SULAMA BAKIM ISCILIK',
                             'SULAMA SISTEMI BAKIM', 'SULAMA SISTEM BAKIM'],

            # Gübre varyasyonları
            'GUBRE': ['GUBRE', 'GUBRELEME', 'GUBRE ATMA', 'GUBRE UYGULAMA', 'GUBRE ISCILIK', 'GÜBRE', 
                     'GÜBRELEME', 'GUBRE ISCI', 'GUBRELEME ISCILIK', 'GUBRE ATMA ISCILIK', 
                     'GUBRE UYGULAMA ISCILIK', 'GUBRE DESTEK', 'GUBRE UYGULAMA DESTEK'],

            # İlaç varyasyonları
            'ILAC': ['ILAC', 'ILACLAMA', 'ILAC UYGULAMA', 'ILAC ISCILIK', 'ILAÇ', 'ILAÇLAMA', 
                    'ILAC ISCI', 'ILACLAMA ISCILIK', 'ILAC UYGULAMA ISCILIK', 'INSECTICIDE', 
                    'HERBICIDE', 'FUNGUCIDE', 'ROUNDUP', 'DUAL', 'TAROT', 'HERBICIDE ROUNDUP',
                    'HERBICIDE DUAL', 'HERBICIDE ROUNDUP DESTEK', 'HERBICIDE DUAL DESTEK',
                    'FUNGUCIDE DESTEK', 'INSECTICIDE DESTEK'],

            # Dikim varyasyonları
            'DIKIM': ['DIKIM', 'DIKME', 'DIKIM ISCILIK', 'DIKIM IŞCILIK', 'DIKIM ISCI', 'DIKME ISCILIK',
                     'DIKIM DESTEK', 'DIKIM ISCILIK DESTEK'],

            # Çapalama varyasyonları
            'CAPALAMA': ['CAPALAMA', 'CAPA', 'CAPALAMA ISCILIK', 'ÇAPALAMA', 'ÇAPALAMA IŞCILIK',
                        'CAPALAMA ISCI', 'ELLE CAPALAMA', 'MEKANIK CAPALAMA', 'FREZELI CAPALAMA',
                        'ELLE CAPALAMA ISCILIK', 'MEKANIK CAPALAMA ISCILIK', 'FREZELI CAPALAMA ISCILIK'],

            # Kiralama varyasyonları
            'KIRALAMA': ['KIRALAMA', 'KIRA', 'TRAKTOR KIRALAMA', 'TRAKTOR KIRA', 'KİRALAMA',
                        'TRAKTÖR KIRALAMA', 'TRAKTÖR KIRA', 'TRAKTOR KIRALAMA ISCILIK',
                        'TRAKTOR KIRA ISCILIK'],

            # İşçilik varyasyonları
            'ISCILIK': ['ISCILIK', 'ISCILIGI', 'ISCI', 'IŞCILIK', 'IŞÇILIK', 'IŞÇI', 'İŞÇİLİK', 
                       'İŞÇİLİĞİ', 'ISCILIG', 'IŞÇILIG', 'ISÇILIK', 'ISÇILIGI', 'ISÇI'],

            # Kırım varyasyonları
            'KIRIM': ['KIRIM', 'KIRMA', 'KIRIM ISCILIK', 'KIRIM ISCI', 'KIRIM DESTEK', 
                     'KIRIM ISCILIK DESTEK', 'KIRMA ISCILIK'],

            # Kurutma varyasyonları
            'KURUTMA': ['KURUTMA', 'KURUTMA ISCILIK', 'KURUTMA ISCI', 'SERA KURUTMA', 
                       'SERA KURUTMA ISCILIK', 'DIKIS MAK', 'DIKIS MAK DESTEK', 
                       'DIKIS MAK ISCILIK', 'SERA KURUTMA KONTROL', 'ISTIFLEME', 'ISTIFLEME ISCILIK'],

            # Kutulama varyasyonları
            'KUTULAMA': ['KUTULAMA', 'KUTULAMA ISCILIK', 'KUTULAMA ISCI', 'TAVLAMA', 'TAVLAMA ISCILIK'],

            # Nakliye varyasyonları
            'NAKLIYE': ['NAKLIYE', 'NAKLIYAT', 'TASIMA', 'MALZEME NAKLIYE', 'MALZEME TASIMA',
                       'MALZEME NAKLIYE ISCILIK', 'MALZEME TASIMA ISCILIK'],

            # Yükleme varyasyonları
            'YUKLEME': ['YUKLEME', 'INDIRME YUKLEME', 'MALZEME YUKLEME', 'MALZEME INDIRME YUKLEME',
                       'YUKLEME ISCILIK', 'INDIRME YUKLEME ISCILIK', 'MALZEME YUKLEME ISCILIK'],

            # Ekipman bakım varyasyonları
            'BAKIM': ['BAKIM', 'TAMIR', 'TAMIRAT', 'BAKIM ISCILIK', 'TAMIR ISCILIK', 'TAMIRAT ISCILIK',
                     'EKIPMAN BAKIM', 'EKIPMAN TAMIR', 'EKIPMAN TAMIRAT', 'EKIPMAN BAKIM TAMIRAT']
        }

    def ai_fuzzy_match(self, input_text, candidates, threshold=0.4):
        """Gelişmiş fuzzy matching ile en iyi eşleşmeyi bulur"""
        try:
            if not input_text or not candidates:
                return None, 0

            if not isinstance(candidates, dict):
                print(f"WARNING - Candidates dict değil: {type(candidates)}")
                return None, 0

            input_normalized = self.normalize_text_advanced(input_text)
            best_match = None
            best_score = 0

            # Fuzzy match sözlüğü oluştur
            fuzzy_dict = self.create_fuzzy_match_dictionary()

            print(f"DEBUG - Normalized input: '{input_normalized}'")

            for candidate_key, candidate_value in candidates.items():
                try:
                    candidate_normalized = self.normalize_text_advanced(str(candidate_key))

                    # 1. Tam eşleşme kontrolü
                    if input_normalized == candidate_normalized:
                        return (candidate_key, candidate_value), 1.0

                    # 2. Fuzzy sözlük kontrolü
                    score_from_dict = 0
                    for standard_term, variations in fuzzy_dict.items():
                        if input_normalized in variations and standard_term in candidate_normalized:
                            score_from_dict = 0.9
                            break
                        elif any(var in input_normalized for var in
                                 variations) and standard_term in candidate_normalized:
                            score_from_dict = 0.8
                            break

                    # 3. Basit fuzzy scoring
                    score1 = simple_fuzzy_ratio(input_normalized, candidate_normalized) / 100.0
                    score2 = simple_partial_ratio(input_normalized, candidate_normalized) / 100.0

                    # 4. Kelime bazlı örtüşme
                    input_words = set(input_normalized.split())
                    candidate_words = set(candidate_normalized.split())

                    if input_words and candidate_words:
                        word_overlap = len(input_words.intersection(candidate_words)) / len(
                            input_words.union(candidate_words))
                    else:
                        word_overlap = 0

                    # 5. Anahtar kelime bonusu
                    keyword_bonus = self.calculate_keyword_bonus(input_normalized, candidate_normalized)

                    # 6. Özel durum bonusu (Aşılama için)
                    special_bonus = 0
                    if 'ASILAMA' in input_normalized and 'ASILAMA' in candidate_normalized:
                        special_bonus = 0.4
                    elif 'ISCILIK' in input_normalized and any(
                            term in candidate_normalized for term in ['ISCILIK', 'IŞCILIK']):
                        special_bonus = 0.3

                    # Final score hesaplama
                    final_score = max(
                        score_from_dict,
                        (
                                score1 * 0.25 + score2 * 0.35 + word_overlap * 0.25 + keyword_bonus * 0.1 + special_bonus * 0.05)
                    )

                    print(
                        f"DEBUG - Candidate '{candidate_key}': skor={final_score:.3f} (dict:{score_from_dict:.2f}, s1:{score1:.2f}, s2:{score2:.2f}, word:{word_overlap:.2f}, bonus:{special_bonus:.2f})")

                    if final_score > best_score and final_score >= threshold:
                        best_score = final_score
                        best_match = (candidate_key, candidate_value)

                except Exception as e:
                    print(f"WARNING - Candidate işlenirken hata: {candidate_key} - {str(e)}")
                    continue

            if best_match:
                print(f"DEBUG - En iyi eşleşme: '{best_match[0]}' -> {best_match[1]} (skor: {best_score:.3f})")
            else:
                print(f"DEBUG - Hiçbir eşleşme bulunamadı (threshold: {threshold})")

            return best_match, best_score

        except Exception as e:
            print(f"ERROR - Enhanced fuzzy match hatası: {str(e)}")
            return None, 0

    def normalize_text_advanced(self, text):
        """EN ÜST SEVİYE gelişmiş metin normalizasyonu - maksimum varyasyon desteği"""
        try:
            if not text:
                return ""

            text = str(text).upper().strip()

            # Genişletilmiş Türkçe karakter dönüşümü
            turkish_chars = {
                'Ç': 'C', 'Ğ': 'G', 'İ': 'I', 'Ö': 'O', 'Ş': 'S', 'Ü': 'U',
                'ç': 'C', 'ğ': 'G', 'ı': 'I', 'ö': 'O', 'ş': 'S', 'ü': 'U',
                'Ž': 'Z', 'ž': 'Z', 'Ñ': 'N', 'ñ': 'N'
            }

            for tr_char, en_char in turkish_chars.items():
                text = text.replace(tr_char, en_char)

            # Yaygın yazım hatalarını düzelt - genişletilmiş liste
            replacements = {
                'ISCILIGI': 'ISCILIK', 'ISÇILIK': 'ISCILIK', 'ISÇILIGI': 'ISCILIK',
                'IŞÇILIK': 'ISCILIK', 'IŞÇILIGI': 'ISCILIK', 'İŞÇİLİK': 'ISCILIK',
                'İŞÇİLİĞİ': 'ISCILIK', 'ISÇILIG': 'ISCILIK', 'IŞÇILIG': 'ISCILIK',
                'TAMIRAT': 'TAMIR', 'TAMIR ISCILIK': 'TAMIR', 'TAMIRAT ISCILIK': 'TAMIR',
                'BAKIM': 'TAMIR', 'BAKIM ISCILIK': 'TAMIR',
                'GUBRELEME': 'GUBRE', 'GUBRE ATMA': 'GUBRE', 'GUBRE UYGULAMA': 'GUBRE',
                'GUBRE ISCILIK': 'GUBRE', 'GUBRELEME ISCILIK': 'GUBRE',
                'ILACLAMA': 'ILAC', 'ILAC UYGULAMA': 'ILAC', 'ILAC ISCILIK': 'ILAC',
                'ILACLAMA ISCILIK': 'ILAC',
                'DIKME': 'DIKIM', 'DIKIM ISCILIK': 'DIKIM', 'DIKME ISCILIK': 'DIKIM',
                'ASIM': 'ASILAMA', 'ASILAMA ISCILIK': 'ASILAMA', 'ASIM ISCILIK': 'ASILAMA',
                'CAPA': 'CAPALAMA', 'CAPALAMA ISCILIK': 'CAPALAMA', 'CAPA ISCILIK': 'CAPALAMA',
                'KIRA': 'KIRALAMA', 'TRAKTOR KIRALAMA': 'KIRALAMA', 'TRAKTOR KIRA': 'KIRALAMA',
                'NAKLIYAT': 'NAKLIYE', 'TASIMA': 'NAKLIYE',
                'SULAMA ISCILIK': 'SULAMA', 'SULAMA SISTEMI ISCILIK': 'SULAMA',
                'SULAMA SISTEM ISCILIK': 'SULAMA', 'SULAMA TAMIR': 'SULAMA TAMIR',
                'SULAMA TAMIRAT': 'SULAMA TAMIR', 'SULAMA BAKIM': 'SULAMA TAMIR',
                'KURUTMA ISCILIK': 'KURUTMA', 'KUTULAMA ISCILIK': 'KUTULAMA',
                'KIRIM ISCILIK': 'KIRIM', 'KIRMA': 'KIRIM',
                'YUKLEME ISCILIK': 'YUKLEME', 'INDIRME YUKLEME': 'YUKLEME',
                'MALZEME YUKLEME': 'YUKLEME', 'MALZEME NAKLIYE': 'NAKLIYE',
                'MEKANIK CAPALAMA': 'MEKANIK CAPALAMA', 'ELLE CAPALAMA': 'ELLE CAPALAMA',
                'FREZELI CAPALAMA': 'MEKANIK CAPALAMA', 'FREZELI': 'MEKANIK',
                'INSECTICIDE': 'ILAC', 'HERBICIDE': 'ILAC', 'FUNGUCIDE': 'ILAC',
                'ROUNDUP': 'HERBICIDE', 'DUAL': 'HERBICIDE', 'TAROT': 'ILAC'
            }
            
            for old, new in replacements.items():
                text = text.replace(old, new)

            # Özel karakterleri temizle (ama bazı önemli karakterleri koru)
            text = re.sub(r'[^\w\s\-_]', ' ', text)

            # Gelişmiş tarımsal terim standartlaştırma - genişletilmiş
            standardizations = {
                # Aşılama varyasyonları
                r'\bASILAMA\b|\bASIM\b|\bASILAMA ISCILIK\b|\bASILAMA ISCILIGI\b': 'ASILAMA',

                # İşçilik varyasyonları
                r'\bISCI\b|\bISCILIK\b|\bISCILIGI\b|\bISCILIG\b': 'ISCILIK',

                # Sulama varyasyonları
                r'\bSULAMA ISCILIK\b|\bSULAMA ISCILIGI\b|\bSULAMA SISTEMI ISCILIK\b': 'SULAMA',
                r'\bSULAMA TAMIR\b|\bSULAMA TAMIRAT\b|\bSULAMA BAKIM\b': 'SULAMA TAMIR',

                # Tamir varyasyonları
                r'\bTAMIR\b|\bTAMIRAT\b|\bBAKIM\b|\bTAMIR ISCILIK\b|\bTAMIRAT ISCILIK\b': 'TAMIR',

                # Gübre varyasyonları
                r'\bGUBRE\b|\bGUBRELEME\b|\bGUBRE ATMA\b|\bGUBRE UYGULAMA\b': 'GUBRE',

                # İlaç varyasyonları
                r'\bILAC\b|\bILACLAMA\b|\bILAC UYGULAMA\b|\bINSECTICIDE\b|\bHERBICIDE\b|\bFUNGUCIDE\b': 'ILAC',

                # Dikim varyasyonları
                r'\bDIKIM\b|\bDIKME\b|\bDIKIM ISCILIK\b': 'DIKIM',

                # Fide varyasyonları
                r'\bFIDE\b|\bFIDELIK\b|\bFIDE CEKIMI\b|\bFIDE CEKIM\b': 'FIDE',

                # Çapalama varyasyonları
                r'\bCAPALAMA\b|\bCAPA\b|\bCAPALAMA ISCILIK\b': 'CAPALAMA',
                r'\bMEKANIK CAPALAMA\b|\bFREZELI CAPALAMA\b': 'MEKANIK CAPALAMA',
                r'\bELLE CAPALAMA\b': 'ELLE CAPALAMA',

                # Kiralama varyasyonları
                r'\bKIRALAMA\b|\bKIRA\b|\bTRAKTOR KIRALAMA\b|\bTRAKTOR KIRA\b': 'KIRALAMA',

                # Nakliye varyasyonları
                r'\bNAKLIYE\b|\bNAKLIYAT\b|\bTASIMA\b': 'NAKLIYE',
                r'\bMALZEME NAKLIYE\b|\bMALZEME TASIMA\b': 'MALZEME NAKLIYE',

                # Yükleme varyasyonları
                r'\bYUKLEME\b|\bINDIRME YUKLEME\b|\bMALZEME YUKLEME\b': 'YUKLEME',

                # Kurutma varyasyonları
                r'\bKURUTMA\b|\bKURUTMA ISCILIK\b|\bSERA KURUTMA\b': 'KURUTMA',

                # Kutulama varyasyonları
                r'\bKUTULAMA\b|\bKUTULAMA ISCILIK\b': 'KUTULAMA',

                # Kırım varyasyonları
                r'\bKIRIM\b|\bKIRMA\b|\bKIRIM ISCILIK\b': 'KIRIM'
            }

            for pattern, replacement in standardizations.items():
                text = re.sub(pattern, replacement, text, flags=re.IGNORECASE)

            # Fazla boşlukları temizle
            text = re.sub(r'\s+', ' ', text).strip()

            return text

        except Exception as e:
            print(f"WARNING - Text normalization hatası: {str(e)}")
            return str(text).upper().strip() if text else ""

    def calculate_keyword_bonus(self, input_text, candidate_text):
        """Anahtar kelime bonusu hesaplar"""
        try:
            critical_keywords = [
                ['SULAMA', 'SU'],
                ['TAMIR', 'TAMIRAT', 'BAKIM'],
                ['GUBRE', 'GUBRELEME'],
                ['ILAC', 'ILACLAMA'],
                ['DIKIM', 'DIKME'],
                ['ASILAMA', 'ASIM'],
                ['KIRALAMA', 'KIRA'],
                ['TRAKTOR', 'TRAKTÖR'],
                ['CAPALAMA', 'CAPA'],
                ['KIRIM', 'KIRMA'],
                ['ISCILIK', 'ISCI']
            ]

            bonus = 0
            for keyword_group in critical_keywords:
                input_has = any(kw in input_text for kw in keyword_group)
                candidate_has = any(kw in candidate_text for kw in keyword_group)
                if input_has and candidate_has:
                    bonus += 0.1

            return min(bonus, 0.5)  # Maximum %50 bonus

        except Exception as e:
            print(f"WARNING - Keyword bonus hesaplarken hata: {str(e)}")
            return 0

    def ai_stage_match(self, safha, operasyon):
        """AI tabanlı stage eşleştirme - Sadece eşleştirme yapar, varsayılan değer çağırana bırakılır"""
        try:
            # Input kontrolü ve temizleme
            safha = str(safha) if safha is not None else ""
            operasyon = str(operasyon) if operasyon is not None else ""
            search_text = f"{safha} {operasyon}".strip()

            print(f"DEBUG - AI Stage arıyor: '{search_text}'")

            # Boş input kontrolü
            if not search_text:
                print("WARNING - Boş search text")
                return None

            # AI fuzzy matching
            if not self.mappings.get('stage_kodlari'):
                print("ERROR - Stage kodları bulunamadı!")
                return None

            best_match, score = self.ai_fuzzy_match(search_text, self.mappings['stage_kodlari'], threshold=0.5)
            if best_match:
                print(f"DEBUG - AI Stage eşleşme: '{best_match[0]}' -> {best_match[1]} (skor: {score:.3f})")
                return best_match[1]

            # Fallback: Basit anahtar kelime eşleştirme
            fallback_hints = {
                'DIKIM': '04', 'ASILAMA': '04', 'FIDELIK': '01', 'FIDE': '01',
                'GUBRE': '03', 'GUBRELEME': '03', 'GÜBRE': '03', 'GÜBRELEME': '03',
                'ILAC': '05', 'ILAÇ': '05', 'ILACLAMA': '05', 'ILAÇLAMA': '05',
                'SULAMA': '06', 'SULAMA SISTEMI': '06',
                'CAPALAMA': '07', 'ÇAPALAMA': '07',
                'KIRIM': '08', 'KIRMA': '08',
                'KURUTMA': '09',
                'KUTULAMA': '10',
                'KIRALAMA': '11', 'KIRA': '11', 'TRAKTOR': '11', 'TRAKTÖR': '11',
                'IS': '11', 'ISCI': '11', 'İŞ': '11', 'İŞÇİ': '11',
                'NAKLIYE': '12',
                'SUPERVISOR': '13',
                'KULTUREL': '14', 'KULTUR': '14', 'KÜLTÜREL': '14', 'KÜLTÜR': '14'
            }

            combined_upper = search_text.upper()
            for hint, kod in fallback_hints.items():
                if hint in combined_upper:
                    print(f"DEBUG - Fallback eşleşme: '{hint}' -> {kod}")
                    return kod

            print(f"DEBUG - Hiçbir stage eşleşmesi bulunamadı: '{search_text}'")
            return None

        except Exception as e:
            print(f"ERROR - Stage match hatası: {str(e)}")
            return None

    def levenshtein_distance(self, s1, s2):
        """Levenshtein distance hesaplar - gelişmiş string benzerliği"""
        if len(s1) < len(s2):
            return self.levenshtein_distance(s2, s1)
        if len(s2) == 0:
            return len(s1)
        
        previous_row = range(len(s2) + 1)
        for i, c1 in enumerate(s1):
            current_row = [i + 1]
            for j, c2 in enumerate(s2):
                insertions = previous_row[j + 1] + 1
                deletions = current_row[j] + 1
                substitutions = previous_row[j] + (c1 != c2)
                current_row.append(min(insertions, deletions, substitutions))
            previous_row = current_row
        
        return previous_row[-1]
    
    def advanced_similarity_score(self, text1, text2):
        """Gelişmiş benzerlik skoru - çoklu algoritma kombinasyonu"""
        try:
            if not text1 or not text2:
                return 0.0
            
            # Normalize et
            t1 = self.normalize_text_advanced(text1)
            t2 = self.normalize_text_advanced(text2)
            
            # 1. Tam eşleşme
            if t1 == t2:
                return 1.0
            
            # 2. Levenshtein distance skoru
            max_len = max(len(t1), len(t2))
            if max_len == 0:
                return 0.0
            lev_dist = self.levenshtein_distance(t1, t2)
            lev_score = 1.0 - (lev_dist / max_len)
            
            # 3. SequenceMatcher skoru
            seq_score = simple_fuzzy_ratio(t1, t2) / 100.0
            
            # 4. Partial ratio skoru
            partial_score = simple_partial_ratio(t1, t2) / 100.0
            
            # 5. Kelime bazlı Jaccard benzerliği
            words1 = set(t1.split())
            words2 = set(t2.split())
            if words1 or words2:
                jaccard = len(words1.intersection(words2)) / len(words1.union(words2)) if words1.union(words2) else 0
            else:
                jaccard = 0
            
            # 6. Ortak karakter oranı
            chars1 = set(t1.replace(' ', ''))
            chars2 = set(t2.replace(' ', ''))
            if chars1 or chars2:
                char_overlap = len(chars1.intersection(chars2)) / len(chars1.union(chars2)) if chars1.union(chars2) else 0
            else:
                char_overlap = 0
            
            # 7. N-gram benzerliği (2-gram)
            def get_ngrams(text, n=2):
                return set([text[i:i+n] for i in range(len(text)-n+1)])
            
            ngrams1 = get_ngrams(t1.replace(' ', ''))
            ngrams2 = get_ngrams(t2.replace(' ', ''))
            if ngrams1 or ngrams2:
                ngram_score = len(ngrams1.intersection(ngrams2)) / len(ngrams1.union(ngrams2)) if ngrams1.union(ngrams2) else 0
            else:
                ngram_score = 0
            
            # 8. Anahtar kelime eşleşme bonusu
            keyword_bonus = self.calculate_keyword_bonus(t1, t2) / 0.5  # Normalize to 0-1
            
            # 9. Önek/sonek eşleşmesi
            prefix_score = 0
            suffix_score = 0
            if len(t1) >= 3 and len(t2) >= 3:
                # İlk 3 karakter
                if t1[:3] == t2[:3]:
                    prefix_score = 0.2
                # Son 3 karakter
                if len(t1) >= 3 and len(t2) >= 3 and t1[-3:] == t2[-3:]:
                    suffix_score = 0.2
            
            # 10. İçerme kontrolü (substring match)
            containment_score = 0
            shorter = t1 if len(t1) <= len(t2) else t2
            longer = t2 if len(t1) <= len(t2) else t1
            if shorter in longer:
                containment_score = len(shorter) / len(longer) if longer else 0
            
            # Ağırlıklı ortalama - en iyi skorları önceliklendir
            final_score = (
                lev_score * 0.20 +
                seq_score * 0.15 +
                partial_score * 0.15 +
                jaccard * 0.15 +
                char_overlap * 0.10 +
                ngram_score * 0.10 +
                keyword_bonus * 0.05 +
                prefix_score * 0.03 +
                suffix_score * 0.02 +
                containment_score * 0.05
            )
            
            return min(final_score, 1.0)  # Maksimum 1.0
            
        except Exception as e:
            print(f"WARNING - Advanced similarity score hatası: {str(e)}")
            return 0.0

    def ai_operation_match(self, safha, operasyon, stage_kodu):
        """AI tabanlı operasyon eşleştirme - EN ÜST SEVİYE GELİŞMİŞ EŞLEŞTİRME"""
        try:
            # Input kontrolü
            if not operasyon:
                print("WARNING - Boş operasyon, varsayılan kullanılıyor")
                return "00", "Bilinmeyen Operasyon", f"{stage_kodu}00"

            search_text = str(operasyon).strip()
            print(f"DEBUG - AI Operasyon arıyor: '{search_text}' (stage: {stage_kodu})")

            # Önce yerel mapping'leri kontrol et
            operations = {}

            # Yerel mapping'lerden stage'e ait operasyonları bul
            if self.mappings and 'operasyonlar' in self.mappings:
                stage_ops = self.mappings['operasyonlar'].get(stage_kodu, {})
                if stage_ops:
                    operations = stage_ops
                    print(f"DEBUG - Yerel mapping'den {len(operations)} operasyon bulundu")

            # Eğer yerelde yoksa API'ye git
            if not operations:
                print(f"DEBUG - Yerelde operasyon bulunamadı, API deneniyor: get_operations_by_stage({stage_kodu})")

                try:
                    # API URL'sini oluştur
                    url = f"{get_api_root()}/get_operations_by_stage/{stage_kodu}"
                    print(f"DEBUG - API URL: {url}")

                    # Thread-safe requests için QNetworkAccessManager kullanılmalı ama
                    # basitçe timeout ile dene
                    import requests
                    response = requests.get(
                        url, timeout=5, headers=merge_auth_headers()
                    )

                    if response.status_code == 200:
                        data = response.json()
                        if data.get('success') and data.get('data'):
                            operations = data['data']
                            print(f"DEBUG - API'den {len(operations)} operasyon alındı")
                        else:
                            print(f"WARNING - API başarısız yanıt: {data}")
                    else:
                        print(f"WARNING - API HTTP {response.status_code} döndürdü")

                except requests.exceptions.Timeout:
                    print("WARNING - API timeout, yerel mapping kullanılıyor")
                except requests.exceptions.ConnectionError:
                    print("WARNING - API bağlantı hatası, yerel mapping kullanılıyor")
                except Exception as e:
                    print(f"WARNING - API çağrısı hatası: {str(e)}")

            # Hala operations yoksa, tüm operasyonlardan stage'e göre filtrele
            if not operations and self.mappings and 'operasyonlar' in self.mappings:
                all_ops = self.mappings['operasyonlar']
                operations = {}
                for stage, ops in all_ops.items():
                    if stage.startswith(stage_kodu):
                        operations.update(ops)
                print(f"DEBUG - Tüm mapping'den {len(operations)} operasyon filtrelendi")

            # Operations kontrolü
            if not operations or not isinstance(operations, dict):
                print(f"WARNING - Geçerli operations verisi yok: {operations}")
                return "00", operasyon, f"{stage_kodu}00"

            # Giriş metnini normalize et
            try:
                normalized_search = self.normalize_text_advanced(search_text)
                print(f"DEBUG - Normalize edilmiş arama metni: {normalized_search}")
            except Exception as e:
                print(f"ERROR - Text normalizasyon hatası: {str(e)}")
                normalized_search = search_text.upper()

            # Yüksek öncelikli özel durumlar ve kesin eşleşmeler
            try:
                # 1. Stage 05 için operasyon kodu her zaman 04 olmalı (INSECTICIDE)
                if stage_kodu == '05':
                    for op_code, op_name in operations.items():
                        if op_code == '04':  # 0504 operasyonu (Insecticide)
                            stage_op_combo = f"{stage_kodu}{op_code}"
                            print(f"DEBUG - Stage 05 kuralı uygulandı: '{op_name}' ({op_code}) -> {stage_op_combo}")
                            return op_code, op_name, stage_op_combo

                # 2. Aşılama özel durumu (Stage 04 için 0403)
                if stage_kodu == '04':
                    try:
                        normalized_safha = self.normalize_text_advanced(str(safha) if safha else "")
                        if 'ASILAMA' in normalized_search or 'ASILAMA' in normalized_safha:
                            for op_code, op_name in operations.items():
                                normalized_op_name = self.normalize_text_advanced(op_name)
                                if op_code == '03' and (
                                        'ASILAMA' in normalized_op_name or 'ASIM' in normalized_op_name):
                                    stage_op_combo = f"{stage_kodu}{op_code}"
                                    print(
                                        f"DEBUG - Aşılama özel durumu uygulandı: '{op_name}' ({op_code}) -> {stage_op_combo}")
                                    return op_code, op_name, stage_op_combo
                    except Exception as e:
                        print(f"WARNING - Aşılama özel durumu hatası: {str(e)}")

                # 3. Sulama Tamirat özel durumu (Stage 06 için 0603)
                if stage_kodu == '06':
                    try:
                        normalized_safha = self.normalize_text_advanced(str(safha) if safha else "")
                        if ('TAMIRAT' in normalized_search or 'TAMIR' in normalized_search or
                                'TAMIRAT' in normalized_safha or 'TAMIR' in normalized_safha):
                            for op_code, op_name in operations.items():
                                normalized_op_name = self.normalize_text_advanced(op_name)
                                if op_code == '03' and (
                                        'TAMIRAT' in normalized_op_name or 'TAMIR' in normalized_op_name):
                                    stage_op_combo = f"{stage_kodu}{op_code}"
                                    print(
                                        f"DEBUG - Sulama Tamirat özel durumu uygulandı: '{op_name}' ({op_code}) -> {stage_op_combo}")
                                    return op_code, op_name, stage_op_combo
                    except Exception as e:
                        print(f"WARNING - Sulama tamirat özel durumu hatası: {str(e)}")

                # 4. Kiralama/Traktör özel durumu (Stage 11 için 1103)
                if stage_kodu == '11' and ('KIRALAMA' in normalized_search or 'TRAKTOR' in normalized_search):
                    try:
                        for op_code, op_name in operations.items():
                            normalized_op_name = self.normalize_text_advanced(op_name)
                            if op_code == '03' and (
                                    'KIRALAMA' in normalized_op_name or 'TRAKTOR' in normalized_op_name):
                                stage_op_combo = f"{stage_kodu}{op_code}"
                                print(
                                    f"DEBUG - Kiralama/Traktör özel durumu uygulandı: '{op_name}' ({op_code}) -> {stage_op_combo}")
                                return op_code, op_name, stage_op_combo
                    except Exception as e:
                        print(f"WARNING - Kiralama/Traktör özel durumu hatası: {str(e)}")

                # Özel kural: Frezeli çapalama -> Mekanik çapalama
                if 'FREZELI' in normalized_search and 'CAPALAMA' in normalized_search:
                    try:
                        for op_code, op_name in operations.items():
                            normalized_op = self.normalize_text_advanced(op_name)
                            if 'MEKANIK' in normalized_op and 'CAPALAMA' in normalized_op:
                                stage_op_combo = f"{stage_kodu}{op_code}"
                                print(
                                    f"DEBUG - Özel kural (Frezeli->Mekanik Çapalama) uygulandı: '{op_name}' ({op_code}) -> {stage_op_combo}")
                                return op_code, op_name, stage_op_combo
                    except Exception as e:
                        print(f"WARNING - Frezeli çapalama özel durumu hatası: {str(e)}")
                
                # Özel kural: MALZEME İNDİRME İŞÇİLİĞİ -> 97 kodu
                if ('MALZEME' in normalized_search and 'INDIRME' in normalized_search) or \
                   ('MALZEME' in normalized_search and 'İNDİRME' in normalized_search) or \
                   ('MALZEME İNDİRME İŞÇİLİĞİ' in normalized_search.upper()):
                    try:
                        # 97 kodunu kontrol et
                        if '97' in operations:
                            op_code = '97'
                            op_name = operations.get('97', 'Malzeme İndirme Yükleme')
                            stage_op_combo = f"{stage_kodu}97"
                            print(f"DEBUG - Özel kural (Malzeme İndirme İşçiliği->97) uygulandı: '{op_name}' ({op_code}) -> {stage_op_combo}")
                            return op_code, op_name, stage_op_combo
                    except Exception as e:
                        print(f"WARNING - Malzeme İndirme İşçiliği özel durumu hatası: {str(e)}")

            except Exception as e:
                print(f"WARNING - Özel durumlar kontrolü hatası: {str(e)}")

            # EN ÜST SEVİYE ANA EŞLEŞTİRME ALGORİTMASI
            try:
                best_score = 0
                best_match = None
                all_scores = []  # Tüm skorları kaydet

                # Her operasyon için gelişmiş benzerlik skoru hesapla
                for op_code, op_name in operations.items():
                    try:
                        normalized_op = self.normalize_text_advanced(op_name)
                        print(f"DEBUG - Karşılaştırma: '{normalized_search}' vs '{normalized_op}'")

                        # 1. Gelişmiş similarity score kullan
                        base_score = self.advanced_similarity_score(search_text, op_name)
                        
                        # 2. Tam eşleşme kontrolü (maksimum skor)
                        if normalized_search == normalized_op:
                            score = 1.0
                            print(f"DEBUG - Tam eşleşme bulundu: {op_name} (skor: {score:.3f})")
                            best_score = score
                            best_match = (op_code, op_name)
                            break  # Tam eşleşme varsa hemen dön
                        
                        # 3. Kelime bazlı gelişmiş analiz
                        search_words = set(normalized_search.split())
                        op_words = set(normalized_op.split())
                        
                        # Ortak kelime skoru (Jaccard)
                        if search_words or op_words:
                            common_words = search_words.intersection(op_words)
                            union_words = search_words.union(op_words)
                            word_jaccard = len(common_words) / len(union_words) if union_words else 0
                        else:
                            word_jaccard = 0
                        
                        # 4. Gelişmiş anahtar kelime eşleşmesi
                        keyword_groups = {
                            'SULAMA': ['SULAMA', 'SU', 'SULAMA SISTEMI', 'SULAMA SISTEM'],
                            'TAMIR': ['TAMIR', 'TAMIRAT', 'BAKIM', 'TAMIR ISCILIK', 'TAMIRAT ISCILIK'],
                            'GUBRE': ['GUBRE', 'GUBRELEME', 'GUBRE ATMA', 'GUBRE UYGULAMA', 'GUBRE ISCILIK'],
                            'ILAC': ['ILAC', 'ILACLAMA', 'ILAC UYGULAMA', 'ILAC ISCILIK', 'INSECTICIDE', 'HERBICIDE', 'FUNGUCIDE'],
                            'DIKIM': ['DIKIM', 'DIKME', 'DIKIM ISCILIK'],
                            'ASILAMA': ['ASILAMA', 'ASIM', 'ASILAMA ISCILIK'],
                            'CAPALAMA': ['CAPALAMA', 'CAPA', 'CAPALAMA ISCILIK', 'MEKANIK CAPALAMA', 'ELLE CAPALAMA'],
                            'KIRALAMA': ['KIRALAMA', 'KIRA', 'TRAKTOR KIRALAMA', 'TRAKTOR KIRA'],
                            'NAKLIYE': ['NAKLIYE', 'NAKLIYAT', 'TASIMA', 'MALZEME NAKLIYE'],
                            'YUKLEME': ['YUKLEME', 'YUKLEME ISCILIK', 'INDIRME YUKLEME', 'MALZEME YUKLEME'],
                            'KURUTMA': ['KURUTMA', 'KURUTMA ISCILIK', 'SERA KURUTMA'],
                            'KUTULAMA': ['KUTULAMA', 'KUTULAMA ISCILIK'],
                            'KIRIM': ['KIRIM', 'KIRMA', 'KIRIM ISCILIK'],
                            'ISCILIK': ['ISCILIK', 'ISCI', 'ISCILIGI', 'ISÇILIK', 'ISÇI', 'ISÇILIGI']
                        }
                        
                        keyword_match_score = 0
                        for group_name, keywords in keyword_groups.items():
                            search_has = any(kw in normalized_search for kw in keywords)
                            op_has = any(kw in normalized_op for kw in keywords)
                            if search_has and op_has:
                                keyword_match_score += 0.15
                        
                        keyword_match_score = min(keyword_match_score, 0.6)  # Maksimum %60
                        
                        # 5. Özel durum bonusları
                        special_bonus = 0
                        
                        # Aşılama özel durumu
                        if 'ASILAMA' in normalized_search and 'ASILAMA' in normalized_op:
                            special_bonus += 0.3
                        
                        # Sulama özel durumu
                        if 'SULAMA' in normalized_search and 'SULAMA' in normalized_op:
                            special_bonus += 0.2
                        
                        # Tamir özel durumu
                        if any(kw in normalized_search for kw in ['TAMIR', 'TAMIRAT', 'BAKIM']) and \
                           any(kw in normalized_op for kw in ['TAMIR', 'TAMIRAT', 'BAKIM']):
                            special_bonus += 0.25
                        
                        # İlaç özel durumu
                        if any(kw in normalized_search for kw in ['ILAC', 'INSECTICIDE', 'HERBICIDE', 'FUNGUCIDE']) and \
                           any(kw in normalized_op for kw in ['ILAC', 'INSECTICIDE', 'HERBICIDE', 'FUNGUCIDE']):
                            special_bonus += 0.2
                        
                        # 6. Pozisyon bazlı eşleşme (başta/sonda)
                        position_bonus = 0
                        if normalized_search.split()[0] == normalized_op.split()[0] if normalized_search.split() and normalized_op.split() else False:
                            position_bonus += 0.1
                        if normalized_search.split()[-1] == normalized_op.split()[-1] if normalized_search.split() and normalized_op.split() else False:
                            position_bonus += 0.1
                        
                        # 7. Uzunluk benzerliği bonusu
                        length_ratio = min(len(normalized_search), len(normalized_op)) / max(len(normalized_search), len(normalized_op)) if max(len(normalized_search), len(normalized_op)) > 0 else 0
                        length_bonus = length_ratio * 0.1
                        
                        # 8. Final skor hesaplama - çoklu algoritma kombinasyonu
                        score = (
                            base_score * 0.40 +           # Gelişmiş similarity (en önemli)
                            word_jaccard * 0.20 +          # Kelime bazlı Jaccard
                            keyword_match_score * 0.20 +   # Anahtar kelime eşleşmesi
                            special_bonus * 0.10 +         # Özel durum bonusları
                            position_bonus * 0.05 +        # Pozisyon bonusu
                            length_bonus * 0.05            # Uzunluk benzerliği
                        )
                        
                        # Skoru normalize et (0-1 arası)
                        score = min(score, 1.0)
                        
                        all_scores.append((op_code, op_name, score))
                        print(f"DEBUG - Operasyon '{op_name}' ({op_code}) için detaylı skor: {score:.4f} "
                              f"(base:{base_score:.3f}, word:{word_jaccard:.3f}, keyword:{keyword_match_score:.3f}, "
                              f"special:{special_bonus:.3f}, pos:{position_bonus:.3f}, len:{length_bonus:.3f})")

                        if score > best_score:
                            best_score = score
                            best_match = (op_code, op_name)

                    except Exception as op_error:
                        print(f"WARNING - Operasyon '{op_code}' işleme hatası: {str(op_error)}")
                        continue

                # Tüm skorları sırala ve logla (debug için)
                if all_scores:
                    all_scores.sort(key=lambda x: x[2], reverse=True)
                    print(f"DEBUG - Top 5 eşleşme:")
                    for i, (code, name, scr) in enumerate(all_scores[:5], 1):
                        print(f"  {i}. {name} ({code}): {scr:.4f}")

                # Eşik değeri kontrolü - daha düşük eşik (daha esnek)
                threshold = 0.12  # %12 eşik değeri (önceden %15)
                if best_match and best_score >= threshold:
                    stage_op_combo = f"{stage_kodu}{best_match[0]}"
                    print(
                        f"DEBUG - ✅ EN İYİ EŞLEŞME: '{best_match[0]}' -> '{best_match[1]}' -> {stage_op_combo} (skor: {best_score:.4f})")
                    return best_match[0], best_match[1], stage_op_combo
                elif best_match:
                    print(f"DEBUG - ⚠️ Eşleşme bulundu ama eşik değerinin altında: {best_score:.4f} < {threshold}")

            except Exception as e:
                print(f"ERROR - Ana eşleştirme algoritması hatası: {str(e)}")

            # Hiçbir eşleşme bulunamadı
            print(f"DEBUG - Hiçbir operasyon eşleşmesi bulunamadı: '{search_text}' - Varsayılan format kullanılıyor")
            return "00", operasyon, f"{stage_kodu}00"

        except Exception as e:
            print(f"ERROR - Operation match genel hatası: {str(e)}")
            print(f"ERROR - Hata tipi: {type(e).__name__}")
            import traceback
            print(f"ERROR - Traceback: {traceback.format_exc()}")
            return "00", operasyon, f"{stage_kodu}00"

    def process_data(self, df):
        """Verileri güvenli şekilde işler"""
        try:
            print(f"DEBUG - Veri işleme başlıyor: {len(df)} satır")

            # Sütun isimlerini güvenli şekilde standartlaştır
            column_mapping = {
                'Ürün Miktar': 'Miktar',
                'Bilgi': 'Miktar Bilgi',
                'Üreti Tarihi': 'Tarih',
            }

            print(f"DEBUG - Excel okuduktan sonraki sütunlar: {df.columns.tolist()}")

            for old_name, new_name in column_mapping.items():
                if old_name in df.columns:
                    df = df.rename(columns={old_name: new_name})

            print(f"DEBUG - Yeniden adlandırma sonrası sütunlar: {df.columns.tolist()}")
            df.columns = df.columns.str.strip()
            df.columns = df.columns.str.strip().str.replace(r'\s+', ' ', regex=True)

            # Gerekli sütunları kontrol et (internal names)
            required_columns = ['Miktar', 'Alt Saha', 'Safha', 'Operasyon', 'Isgücü Tipi', 'Miktar Bilgi', 'Tarih']
            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                error_msg = f"Gerekli sütunlar eksik: {missing_columns}"
                print(f"ERROR - {error_msg}")
                raise ValueError(error_msg)

            # Veri temizliği
            df['Miktar'] = pd.to_numeric(df['Miktar'], errors='coerce').fillna(0)
            df['Alt Saha'] = df['Alt Saha'].fillna('').astype(str).str.strip().str.upper()
            df['Safha'] = df['Safha'].fillna('').astype(str).str.strip().str.upper()
            df['Operasyon'] = df['Operasyon'].fillna('').astype(str).str.strip().str.upper()
            df['Isgücü Tipi'] = df['Isgücü Tipi'].fillna('').astype(str).str.strip().str.upper()

            # BÖLGE KODU eşleştirme - API'den alınan verilerle
            print("DEBUG - Bölge kodları API'den alınıyor...")
            if self.mappings.get('bolge_kodlari'):
                print(f"DEBUG - API'den alınan bölge kodları: {self.mappings['bolge_kodlari']}")
                
                # Bölge kodlarını özel eşleştirme fonksiyonu ile eşleştir
                df['BÖLGE KODU'] = df['Alt Saha'].apply(self.get_region_code_from_api)
                
                # Eşleşme sonuçlarını kontrol et
                matched_count = df['BÖLGE KODU'].value_counts()
                print(f"DEBUG - Bölge kodu eşleşme sonuçları: {matched_count.to_dict()}")
                
                # Eşleşmeyen değerleri logla
                unmatched = df[df['BÖLGE KODU'] == '22']['Alt Saha'].unique()
                if len(unmatched) > 0:
                    print(f"DEBUG - Eşleşmeyen Alt Saha değerleri: {unmatched}")
            else:
                print("WARNING - API'den bölge kodları alınamadı, varsayılan değer kullanılıyor")
                df['BÖLGE KODU'] = '22'

            # Boş miktar satırlarını filtrele
            df = df[df['Miktar'] > 0]
            print(f"DEBUG - Filtrelenmiş veri: {len(df)} satır")

            if df.empty:
                raise ValueError("Miktar > 0 olan satır bulunamadı!")

            # Her satırı güvenli şekilde işle
            output_data = []
            for idx, row in df.iterrows():
                try:
                    processed_row = self.process_row(row)
                    if processed_row:
                        output_data.append(processed_row)
                        if idx % 10 == 0:  # Her 10 satırda bir progress güncelle
                            progress = 30 + int((idx / len(df)) * 50)
                            self.progress_updated.emit(min(progress, 79))
                except Exception as e:
                    print(f"WARNING - Satır {idx} işlenirken hata: {str(e)}")
                    continue

            if not output_data:
                raise ValueError("Hiçbir satır başarıyla işlenemedi!")

            # DataFrame oluştur
            processed_df = pd.DataFrame(output_data)
            print(f"DEBUG - İşlenmiş veri: {len(processed_df)} satır")

            # Eksik sütunları kontrol et ve ekle
            required_output_columns = [
                "KAYNAK TİPİ KODU", "STAGE KODU", "STAGE-OPERASYON KODU",
                "Safha", "Operasyon", "Birim", "BÖLGE KODU", "Miktar", "Toplam", "Açıklama"
            ]

            for col in required_output_columns:
                if col not in processed_df.columns:
                    if col in ["Miktar", "Toplam"]:
                        processed_df[col] = 0.00
                    else:
                        processed_df[col] = ""

            return processed_df

        except Exception as e:
            print(f"ERROR - Veri işleme hatası: {str(e)}")
            return None

    def process_row(self, row):
        """Tek satırı güvenli şekilde işler"""
        try:
            # Güvenli veri çıkarma
            safha = str(row.get('Safha', '')).strip() if pd.notna(row.get('Safha')) else ''
            operasyon = str(row.get('Operasyon', '')).strip() if pd.notna(row.get('Operasyon')) else ''
            isgucu_tipi = str(row.get('Isgücü Tipi', '')).strip() if pd.notna(row.get('Isgücü Tipi')) else ''
            miktar_bilgi = str(row.get('Miktar Bilgi', '')).strip() if pd.notna(row.get('Miktar Bilgi')) else ''
            miktar = float(row.get('Miktar', 0))

            if miktar <= 0:
                return None

            # KAYNAK TİPİ KODU belirleme
            kaynak_tipi_kodu = self.determine_kaynak_tipi_kodu(isgucu_tipi, operasyon, safha)

            # AI tabanlı STAGE KODU belirleme
            stage_kodu = self.ai_stage_match(safha, operasyon)
            if not stage_kodu:
                stage_kodu = '11'  # Varsayılan

            # AI tabanlı STAGE-OPERASYON KODU ve Harcama Kalemi belirleme
            operasyon_kodu, harcama_kalemi, stage_operasyon_kodu = self.ai_operation_match(safha, operasyon, stage_kodu)
            if not stage_operasyon_kodu:
                stage_operasyon_kodu = f"{stage_kodu}00"
                harcama_kalemi = operasyon

            # Birim belirleme
            birim = self.determine_birim(operasyon, isgucu_tipi, miktar_bilgi)

            # Birim ücret hesaplama
            birim_ucret = 1200.00  # Varsayılan
            if self.mappings.get('birim_ucretler') and birim in self.mappings['birim_ucretler']:
                birim_ucret = float(self.mappings['birim_ucretler'][birim])

            toplam = miktar * birim_ucret

            # Açıklama - Safha ve Harcama Kalemi'nden mantıksal olarak oluştur
            if miktar_bilgi:
                aciklama = miktar_bilgi
            else:
                # Safha ve Harcama Kalemi'ni mantıksal olarak birleştir
                if safha and harcama_kalemi:
                    # Her ikisi de varsa: "Safha - Harcama Kalemi" formatında
                    aciklama = f"{safha} - {harcama_kalemi}"
                elif safha:
                    # Sadece safha varsa: "Safha işlemi" formatında
                    aciklama = f"{safha} işlemi"
                elif harcama_kalemi:
                    # Sadece harcama kalemi varsa: "Harcama Kalemi harcaması" formatında
                    aciklama = f"{harcama_kalemi} harcaması"
                else:
                    # Hiçbiri yoksa: "Genel harcama"
                    aciklama = "Genel harcama"

            # BÖLGE KODU - API'den alınan verilerle
            bolge_kodu = '22'  # Varsayılan değer
            try:
                # Önce row'dan gelen değeri kontrol et
                if 'BÖLGE KODU' in row and pd.notna(row['BÖLGE KODU']):
                    bolge_kodu = str(row['BÖLGE KODU']).strip()
                    print(f"DEBUG - Satır bölge kodu: {bolge_kodu}")
                else:
                    # Eğer row'da yoksa, Alt Saha'dan API mapping'i kullanarak belirle
                    alt_saha = str(row.get('Alt Saha', '')).strip()
                    if alt_saha:
                        bolge_kodu = self.get_region_code_from_api(alt_saha)
                        print(f"DEBUG - API'den bölge kodu belirlendi: '{alt_saha}' -> '{bolge_kodu}'")
                    else:
                        print("DEBUG - Alt Saha bilgisi yok, varsayılan bölge kodu kullanılıyor")
            except Exception as e:
                print(f"WARNING - Bölge kodu belirleme hatası: {str(e)}, varsayılan değer kullanılıyor")

            return {
                "KAYNAK TİPİ KODU": kaynak_tipi_kodu,
                "STAGE KODU": stage_kodu,
                "STAGE-OPERASYON KODU": stage_operasyon_kodu,
                "Safha": safha,
                "Operasyon": operasyon,
                "Harcama Kalemi": harcama_kalemi if harcama_kalemi else operasyon,
                "Birim": birim,
                "Miktar": miktar,
                "Toplam": toplam,
                "Açıklama": aciklama,
                "BÖLGE KODU": bolge_kodu
            }

        except Exception as e:
            print(f"ERROR - Satır işleme hatası: {str(e)}")
            return None

    def determine_kaynak_tipi_kodu(self, isgucu_tipi, operasyon, safha):
        """Kaynak tipi kodunu belirler"""
        try:
            # İşgücü tipi kontrolü
            if any(keyword in isgucu_tipi.upper() for keyword in ['ERKEK', 'KADIN']):
                return '01'

            # Operasyon içeriği kontrolü
            text_to_check = f"{operasyon} {safha}".upper()

            if 'KIRALAMA' in text_to_check or 'KIRA' in text_to_check:
                return '05'
            elif any(keyword in text_to_check for keyword in ['MALZEME', 'GÜBRE', 'İLAÇ']):
                return '02'
            elif 'HİZMET' in text_to_check:
                return '03'
            elif 'ENERJİ' in text_to_check:
                return '04'

            return '01'  # Varsayılan işçilik

        except Exception as e:
            print(f"WARNING - Kaynak tipi kodu belirleme hatası: {str(e)}")
            return '01'

    def determine_birim(self, operasyon, isgucu_tipi, miktar_bilgi):
        """Birim belirler"""
        try:
            operasyon = operasyon.upper()
            isgucu_tipi = isgucu_tipi.upper()

            # İşgücü içeren her şey Yevmiye
            if any(x in isgucu_tipi for x in ['ERKEK', 'KADIN', 'ISÇI', 'IŞÇI']):
                return 'YEVMİYE'

            # Traktör ve kiralama işleri
            if any(x in operasyon for x in ['TRAKTÖR', 'KİRALAMA']):
                return 'ADET'

            return 'YEVMİYE'  # Varsayılan

        except Exception as e:
            print(f"WARNING - Birim belirleme hatası: {str(e)}")
            return 'YEVMİYE'

    def get_region_code_from_api(self, alt_saha):
        """Alt Saha'dan API'den bölge kodunu alır"""
        try:
            if not alt_saha or not self.mappings.get('bolge_kodlari'):
                return '22'  # Varsayılan
            
            alt_saha_clean = str(alt_saha).strip().upper()
            print(f"DEBUG - Alt Saha aranıyor: '{alt_saha_clean}'")
            print(f"DEBUG - Mevcut bölge kodları: {list(self.mappings['bolge_kodlari'].keys())}")
            
            # Özel durumlar için manuel eşleştirme (ÖNCE KONTROL ET)
            special_mappings = {
                'MAN-PMI SCV': '23',
                'MAN PMI SCV': '23',
                'MAN PMI': '23',
                'MNS-PMI SCV': '23',  # Eski format desteği
                'MNS PMI SCV': '23',  # Eski format desteği
                'MNS PMI': '23',      # Eski format desteği
                'PMI SCV': '23'
            }
            
            for pattern, kod in special_mappings.items():
                if pattern.upper() in alt_saha_clean:
                    print(f"DEBUG - Özel eşleşme bulundu: '{alt_saha_clean}' -> '{pattern}' ({kod})")
                    return kod
            
            # PMI SCV için özel kontrol (sadece PMI SCV içeren değerler için)
            if 'PMI' in alt_saha_clean and 'SCV' in alt_saha_clean:
                # MAN ile başlayan PMI SCV'ler için
                if 'MAN' in alt_saha_clean:
                    print(f"DEBUG - MAN PMI SCV özel eşleşmesi: '{alt_saha_clean}' -> 23")
                    return '23'
                # MNS ile başlayan PMI SCV'ler için (eski format desteği)
                elif 'MNS' in alt_saha_clean:
                    print(f"DEBUG - MNS PMI SCV özel eşleşmesi: '{alt_saha_clean}' -> 23")
                    return '23'
                # Diğer PMI SCV'ler için varsayılan
                else:
                    print(f"DEBUG - PMI SCV özel eşleşmesi: '{alt_saha_clean}' -> 16")
                    return '16'
            
            # Tam eşleşme kontrolü (büyük/küçük harf duyarsız)
            for kod, ad in self.mappings['bolge_kodlari'].items():
                if alt_saha_clean == ad.upper():
                    print(f"DEBUG - Tam eşleşme bulundu: '{alt_saha_clean}' -> '{ad}' ({kod})")
                    return kod
            
            # Kısmi eşleşme kontrolü - Alt Saha değeri bölge adının içinde var mı?
            for kod, ad in self.mappings['bolge_kodlari'].items():
                ad_upper = ad.upper()
                if alt_saha_clean in ad_upper:
                    print(f"DEBUG - Kısmi eşleşme bulundu (Alt Saha içinde): '{alt_saha_clean}' -> '{ad}' ({kod})")
                    return kod
                # Ters kontrol - bölge adı Alt Saha içinde var mı?
                if ad_upper in alt_saha_clean:
                    print(f"DEBUG - Kısmi eşleşme bulundu (Bölge adı içinde): '{alt_saha_clean}' -> '{ad}' ({kod})")
                    return kod
            
            # Kelime bazlı eşleşme kontrolü
            for kod, ad in self.mappings['bolge_kodlari'].items():
                ad_upper = ad.upper()
                # Ortak kelime kontrolü
                alt_words = set(alt_saha_clean.split())
                ad_words = set(ad_upper.split())
                common_words = alt_words.intersection(ad_words)
                if common_words and len(common_words) >= 2:  # En az 2 ortak kelime
                    print(f"DEBUG - Kelime eşleşmesi bulundu: '{alt_saha_clean}' -> '{ad}' ({kod}) - Ortak kelimeler: {common_words}")
                    return kod
            
            print(f"DEBUG - '{alt_saha_clean}' için bölge kodu bulunamadı, varsayılan kullanılıyor")
            return '22'
            
        except Exception as e:
            print(f"WARNING - Bölge kodu API'den alınırken hata: {str(e)}")
            return '22'

    def group_data(self, df):
        """Verileri gruplar"""
        try:
            # Operasyon isimlerini standartlaştır
            if 'Operasyon' in df.columns:
                df['Operasyon'] = df['Operasyon'].str.upper().replace({
                    'SULAMA ISÇILIGI': 'SULAMA',
                    'SULAMA ISÇILIĞI': 'SULAMA',
                    'SULAMA SISTEMI TAMIR ISÇILIGI': 'SULAMA TAMIRAT',
                    'ILAÇ UYGULAMA ISÇILIK': 'ILAÇ UYGULAMA',
                    'GÜBRE ATMA ISÇILIK': 'GÜBRE UYGULAMA',
                    'ASILAMA ISÇILIK': 'AŞILAMA',
                    'FIDE ÇEKIMI': 'FIDE ÇEKİMİ'
                })

            # Ana gruplama alanları
            grouping_columns = [
                "KAYNAK TİPİ KODU", "STAGE KODU", "STAGE-OPERASYON KODU",
                "Safha", "Birim", "BÖLGE KODU"
            ]

            # Operasyon sütunu varsa gruplamaya ekle
            if 'Operasyon' in df.columns:
                grouping_columns.append("Operasyon")

            grouped_df = df.groupby(grouping_columns).agg({
                "Miktar": "sum",
                "Toplam": "sum",
                "Açıklama": lambda x: ", ".join(x.dropna().astype(str).unique()[:3])  # İlk 3 açıklama
            }).reset_index()

            # Harcama kalemi belirle
            if 'Operasyon' in grouped_df.columns:
                grouped_df["Harcama Kalemi"] = grouped_df["Operasyon"]  # Operasyon bazlı harcama kalemi
            else:
                grouped_df["Harcama Kalemi"] = grouped_df["Safha"]  # Safha bazlı harcama kalemi

            # Birim ücret yeniden hesapla - güvenli şekilde
            birim_ucret_mapping = self.mappings.get('birim_ucretler', {})
            if not birim_ucret_mapping:
                print("WARNING - Birim ücret mapping boş, varsayılan değer kullanılıyor")
                # Varsayılan birim ücretleri
                birim_ucret_mapping = {
                    'YEVMİYE': 1250.00,
                    'ADET': 1500.00,
                    'SAAT': 100.00,
                    'GÜN': 800.00
                }

            grouped_df["Birim ücret"] = grouped_df["Birim"].map(birim_ucret_mapping).fillna(1200.00)

            # Toplamı yeniden hesapla
            grouped_df["Toplam"] = grouped_df["Miktar"] * grouped_df["Birim ücret"]

            # Sıra numarası ve tarih
            grouped_df.insert(0, "No", range(1, len(grouped_df) + 1))
            grouped_df.insert(1, "Tarih", datetime.today().strftime("%d.%m.%Y"))

            # Final sütun sıralaması
            final_columns = [
                "No", "Tarih", "BÖLGE KODU", "KAYNAK TİPİ KODU", "STAGE KODU", "STAGE-OPERASYON KODU",
                "Safha", "Harcama Kalemi", "Birim", "Miktar", "Birim ücret", "Toplam", "Açıklama"
            ]

            # Eksik sütunları ekle
            for col in final_columns:
                if col not in grouped_df.columns:
                    if col in ["Miktar", "Birim ücret", "Toplam"]:
                        grouped_df[col] = 0.00
                    else:
                        grouped_df[col] = ""

            # Final DataFrame'i oluştur
            final_df = grouped_df[final_columns].copy()

            return final_df

        except Exception as e:
            print(f"ERROR - Gruplama hatası: {str(e)}")
            raise
class HarcamaTab(QWidget):
    def __init__(self, api_client, user_id=None):
        super().__init__()
        self.api_client = api_client
        self.user_id = user_id
        self.input_file = None
        self.df_result = None
        self.thread = None

        # Thread'in silinmesini engellemek için attribute ekleyin
        #self.setAttribute(Qt.WA_DeleteOnClose, False)

        # Logging ayarları
        self.setup_logging()
        self.init_ui()

    def cleanup_thread(self):
        """Thread'i güvenli şekilde temizle"""
        if self.thread:
            if hasattr(self.thread, 'stop'):
                self.thread.stop()

            if self.thread.isRunning():
                self.thread.quit()
                self.thread.wait(1000)  # 1 saniye bekle

            self.thread.deleteLater()
            self.thread = None

    def closeEvent(self, event):
        """Pencere kapanırken thread'i temizle"""
        self.cleanup_thread()
        event.accept()
    def setup_logging(self):
        """Logging ayarlarını yapılandır"""
        try:
            # PyInstaller paketinde log dosyası için geçici dizin kullan
            if hasattr(sys, '_MEIPASS'):
                log_dir = os.path.join(os.path.expanduser("~"), "AppData", "Local", "HarcamaMasrafTakip")
            else:
                log_dir = "."

            if not os.path.exists(log_dir):
                os.makedirs(log_dir, exist_ok=True)

            log_file = os.path.join(log_dir, "app.log")

            logging.basicConfig(
                level=logging.INFO,
                format='%(asctime)s - %(levelname)s - %(message)s',
                handlers=[
                    logging.FileHandler(log_file, encoding='utf-8'),
                    logging.StreamHandler()
                ]
            )
            logging.info("Logging başlatıldı")
        except Exception as e:
            print(f"Logging ayarlanırken hata: {e}")

    def get_resource_path(self, relative_path):
        """PyInstaller için kaynak dosyası yolunu döndürür"""
        try:
            # PyInstaller temp klasörü
            base_path = sys._MEIPASS
            logging.info(f"PyInstaller bundle path: {base_path}")
        except Exception:
            # Normal geliştirme ortamı
            base_path = os.path.abspath(".")
            logging.info(f"Development path: {base_path}")

        full_path = os.path.join(base_path, relative_path)
        logging.info(f"Resource path: {full_path}")
        return full_path

    # get_user_file_path metodunu kaldırdık çünkü kullanıcı dosyaları için gerekli değil

    def update_data(self, data=None):
        """Update data from API or use provided data"""
        try:
            logging.info("Data güncelleniyor...")
            # If data is provided, use it directly
            if data:
                if hasattr(self, 'thread') and self.thread:
                    self.thread.mappings = data
                logging.info("Data direkt olarak ayarlandı")
                return True

            # Otherwise get data from API
            all_data = self.api_client.get_all_data()
            if all_data:
                # Update the thread's mappings if it exists
                if hasattr(self, 'thread') and self.thread:
                    self.thread.mappings = all_data
                logging.info("API'den data alındı")
                return True
            logging.warning("API'den data alınamadı")
            return False
        except Exception as e:
            logging.error(f"Error updating data: {str(e)}")
            print(f"Error updating data: {str(e)}")
            return False

    def init_ui(self):
        logging.info("UI başlatılıyor...")
        self.setWindowTitle('Excel İşleme Aracı')
        self.setGeometry(100, 100, 1000, 800)

        self.setStyleSheet("background-color: #f9fafb;")

        layout = QVBoxLayout()

        # Başlık
        title = QLabel('AI Destekli Harcama Sınıfı Dönüştürme Aracı')
        title.setStyleSheet('font-size: 18px; font-weight: bold; color: #1e293b;')
        title.setAlignment(Qt.AlignCenter)
        layout.addWidget(title)

        # AI Info Label
        ai_info = QLabel('🤖 AI tabanlı fuzzy matching ile %100 doğru eşleştirme')
        ai_info.setStyleSheet('font-size: 12px; color: #475569; font-style: italic;')
        ai_info.setAlignment(Qt.AlignCenter)
        layout.addWidget(ai_info)

        # Girdi dosyası seçimi
        input_layout = QHBoxLayout()
        self.input_file_label = QLabel('Girdi Excel Dosyası: Seçilmedi')
        self.input_file_label.setStyleSheet("color: #1e293b; font-size: 14px;")
        input_btn = QPushButton('Dosya Seç')
        input_btn.clicked.connect(self.select_input_file)
        input_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:0.5 #2563eb, stop:1 #1d4ed8);
                color: white;
                font-size: 16px;
                border-radius: 8px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2563eb, stop:0.5 #1d4ed8, stop:1 #1e40af);
            }
        """)
        input_layout.addWidget(self.input_file_label)
        input_layout.addWidget(input_btn)
        layout.addLayout(input_layout)

        # İşle Butonu
        self.process_btn = QPushButton('🚀 AI ile İşle')
        self.process_btn.clicked.connect(self.process_excel)
        self.process_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:0.5 #2563eb, stop:1 #1d4ed8);
                color: white;
                font-size: 18px;
                font-weight: bold;
                border-radius: 10px;
                padding: 12px 24px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2563eb, stop:0.5 #1d4ed8, stop:1 #1e40af);
            }
        """)
        layout.addWidget(self.process_btn)

        # İlerleme Çubuğu
        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid #e5e7eb;
                border-radius: 8px;
                text-align: center;
                background-color: #ffffff;
                color: #1e293b;
                height: 25px;
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3b82f6, stop:1 #2563eb);
                border-radius: 6px;
            }
        """)
        layout.addWidget(self.progress_bar)

        # Tablo görünümü - Kurumsal tema
        self.table_widget = QTableWidget()
        self.table_widget.setStyleSheet("""
            QTableWidget {
                border: 2px solid #e5e7eb;
                gridline-color: #e5e7eb;
                background-color: #ffffff;
                color: #1e293b;
                border-radius: 8px;
            }
            QTableWidget::item {
                padding: 8px;
                border: none;
            }
            QTableWidget::item:selected {
                background-color: #dbeafe;
                color: #2563eb;
            }
            QHeaderView::section {
                background: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                    stop:0 #3b82f6, stop:1 #2563eb);
                color: white;
                padding: 10px;
                border: none;
                font-weight: bold;
                font-size: 13px;
            }
        """)
        layout.addWidget(self.table_widget)

        # Butonlar layout'u
        buttons_layout = QHBoxLayout()
        
        # Kaydet butonu (başta gizli)
        self.save_db_btn = QPushButton('💾 Kaydet')
        self.save_db_btn.clicked.connect(self.save_to_database)
        self.save_db_btn.hide()
        self.save_db_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #2563eb, stop:0.5 #1d4ed8, stop:1 #1e40af);
                color: white;
                font-size: 16px;
                font-weight: bold;
                border-radius: 8px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #7c3aed, stop:0.5 #6d28d9, stop:1 #5b21b6);
            }
        """)
        buttons_layout.addWidget(self.save_db_btn)
        
        # Manuel Ekle butonu (başta gizli)
        self.manual_add_btn = QPushButton('➕ Manuel Ekle')
        self.manual_add_btn.clicked.connect(self.show_manual_add_dialog)
        self.manual_add_btn.hide()
        self.manual_add_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #f59e0b, stop:0.5 #d97706, stop:1 #b45309);
                color: white;
                font-size: 16px;
                border-radius: 8px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #d97706, stop:0.5 #b45309, stop:1 #92400e);
            }
        """)
        buttons_layout.addWidget(self.manual_add_btn)
        
        # Excel'e Aktar butonu (başta gizli)
        self.save_output_btn = QPushButton('📊 Excel\'e Aktar')
        self.save_output_btn.clicked.connect(self.save_output_file)
        self.save_output_btn.hide()
        self.save_output_btn.setStyleSheet("""
            QPushButton {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #4ade80, stop:0.5 #22c55e, stop:1 #16a34a);
                color: white;
                font-size: 16px;
                border-radius: 8px;
                padding: 10px 20px;
            }
            QPushButton:hover {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #3bcc70, stop:0.5 #1db954, stop:1 #15803d);
            }
        """)
        buttons_layout.addWidget(self.save_output_btn)
        
        layout.addLayout(buttons_layout)

        self.setLayout(layout)
        logging.info("UI başlatıldı")

    def select_input_file(self):
        try:
            logging.info("Dosya seçimi başlatılıyor...")

            # Güvenli başlangıç dizini
            start_dir = os.path.expanduser("~")
            if not os.path.exists(start_dir):
                start_dir = "."

            logging.info(f"Başlangıç dizini: {start_dir}")

            file_name, _ = QFileDialog.getOpenFileName(
                self,
                "Excel Dosyasını Seç",
                start_dir,
                "Excel Files (*.xlsx *.xls);;All Files (*.*)"
            )

            if file_name:
                logging.info(f"Seçilen dosya: {file_name}")

                # Dosyanın varlığını kontrol et
                if not os.path.exists(file_name):
                    raise FileNotFoundError(f"Dosya bulunamadı: {file_name}")

                # Dosyanın okunabilir olup olmadığını kontrol et
                if not os.access(file_name, os.R_OK):
                    raise PermissionError(f"Dosya okunamıyor: {file_name}")

                # Kullanıcı dosyası direkt olarak kullanılır
                self.input_file = file_name
                file_basename = os.path.basename(file_name)
                self.input_file_label.setText(f"Girdi Excel Dosyası: {file_basename}")

                logging.info(f"Dosya başarıyla seçildi: {self.input_file}")

        except Exception as e:
            error_msg = f"Dosya seçim hatası: {str(e)}"
            logging.error(error_msg)
            print(error_msg)
            QMessageBox.critical(self, "Hata", error_msg)

    def process_excel(self):
        try:
            logging.info("Excel işleme başlatılıyor...")

            if not self.input_file:
                QMessageBox.warning(self, "Uyarı", "Lütfen bir giriş dosyası seçin.")
                return

            if not os.path.exists(self.input_file):
                error_msg = f"Dosya bulunamadı: {self.input_file}"
                logging.error(error_msg)
                QMessageBox.critical(self, "Hata", error_msg)
                return

            # Mevcut thread varsa temizle
            if self.thread:
                if self.thread.isRunning():
                    self.thread.requestInterruption()
                    self.thread.quit()
                    self.thread.wait(2000)  # 2 saniye bekle
                self.thread.deleteLater()
                self.thread = None

            self.process_btn.setEnabled(False)
            self.progress_bar.setValue(0)
            self.table_widget.clear()
            self.save_output_btn.hide()
            self.df_result = None

            logging.info("Thread başlatılıyor...")

            # Thread'i oluştur
            self.thread = ExcelProcessorThread(self.input_file, self.api_client)

            # Sinyal bağlantılarını kur
            try:
                self.thread.progress_updated.connect(self.progress_bar.setValue, Qt.QueuedConnection)
                print("DEBUG - progress_updated bağlandı (QueuedConnection)")
            except Exception as e:
                print(f"DEBUG - progress_updated bağlantı hatası: {e}")
                
            try:
                # processing_finished sinyalini main thread'de çalışacak şekilde bağla
                self.thread.processing_finished.connect(self.display_result, Qt.QueuedConnection)
                print("DEBUG - processing_finished bağlandı (QueuedConnection)")
            except Exception as e:
                print(f"DEBUG - processing_finished bağlantı hatası: {e}")
                # Alternatif bağlantı dene
                try:
                    self.thread.processing_finished.connect(self.display_result)
                    print("DEBUG - processing_finished alternatif bağlantı başarılı")
                except Exception as e2:
                    print(f"DEBUG - Alternatif bağlantı da başarısız: {e2}")
                    # Son çare: doğrudan bağlantı
                    try:
                        self.thread.processing_finished.connect(self.display_result, Qt.DirectConnection)
                        print("DEBUG - processing_finished DirectConnection başarılı")
                    except Exception as e3:
                        print(f"DEBUG - DirectConnection da başarısız: {e3}")
                        # En son çare: lambda ile bağla
                        try:
                            self.thread.processing_finished.connect(lambda df: self.display_result(df))
                            print("DEBUG - processing_finished lambda bağlantısı başarılı")
                        except Exception as e4:
                            print(f"DEBUG - Lambda bağlantısı da başarısız: {e4}")
                
            try:
                self.thread.error_occurred.connect(self.show_error, Qt.QueuedConnection)
                print("DEBUG - error_occurred bağlandı (QueuedConnection)")
            except Exception as e:
                print(f"DEBUG - error_occurred bağlantı hatası: {e}")
                
            try:
                self.thread.finished.connect(self.on_thread_finished, Qt.QueuedConnection)
                print("DEBUG - finished bağlandı (QueuedConnection)")
            except Exception as e:
                print(f"DEBUG - finished bağlantı hatası: {e}")

            # Bağlantıları kontrol et ve logla
            logging.info("Sinyal bağlantıları kuruldu:")
            logging.info("- progress_updated: Bağlandı")
            logging.info("- processing_finished: Bağlandı")
            logging.info("- error_occurred: Bağlandı")
            logging.info("- finished: Bağlandı")

            # Sinyal bağlantılarını test et
            print("DEBUG - Sinyal bağlantıları test ediliyor...")
            try:
                # Test sinyali gönder
                self.thread.progress_updated.emit(0)
                print("DEBUG - progress_updated sinyali test edildi")
            except Exception as e:
                print(f"DEBUG - progress_updated sinyal test hatası: {e}")

            # Thread'i başlat
            self.thread.start()
            logging.info("Thread başarıyla başlatıldı")
            print("DEBUG - Thread start() çağrıldı")
            
            # Thread'in başlatıldığını kontrol et
            if self.thread.isRunning():
                print("DEBUG - Thread başarıyla çalışıyor")
            else:
                print("DEBUG - Thread başlatılamadı!")

        except Exception as e:
            error_msg = f"Genel işlem hatası: {str(e)}"
            logging.error(error_msg)
            print(f"DEBUG - Exception detayı: {type(e).__name__}: {str(e)}")
            import traceback
            print(f"DEBUG - Traceback: {traceback.format_exc()}")
            QMessageBox.critical(self, "Hata", f"İşlem başlatılamadı: {str(e)}")
            self.process_btn.setEnabled(True)
    def on_thread_finished(self):
        """Thread tamamen bittiğinde çağrılır"""
        logging.info("Thread çalışması tamamlandı")
        print("DEBUG - on_thread_finished çağrıldı")
        
        # Thread'in durumunu kontrol et
        if self.thread:
            logging.info(f"Thread durumu: isRunning={self.thread.isRunning()}, isFinished={self.thread.isFinished()}")
            
            # Thread'i temizle
            self.thread.deleteLater()
            self.thread = None
            logging.info("Thread temizlendi")
        
        # UI'yi güncelle
        self.process_btn.setEnabled(True)
        logging.info("Process butonu tekrar aktif edildi")
        
        # Eğer display_result çağrılmadıysa, manuel olarak kontrol et
        if not hasattr(self, 'df_result') or self.df_result is None:
            print("DEBUG - display_result çağrılmamış, manuel kontrol yapılıyor")
            print("DEBUG - Thread'den sonuç alınmaya çalışılıyor...")
            
            # Thread'den sonucu almaya çalış
            if hasattr(self, 'thread') and self.thread:
                try:
                    # Thread'in sonucunu kontrol et
                    print(f"DEBUG - Thread durumu: isRunning={self.thread.isRunning()}, isFinished={self.thread.isFinished()}")
                    
                    # Eğer thread hala çalışıyorsa bekle
                    if self.thread.isRunning():
                        print("DEBUG - Thread hala çalışıyor, bekleniyor...")
                        self.thread.wait(5000)  # 5 saniye bekle
                    
                    print("DEBUG - Thread tamamen bitti")
                    
                    # Manuel olarak UI'yi güncelle
                    print("DEBUG - Manuel UI güncelleme yapılıyor...")
                    self.process_btn.setEnabled(True)
                    print("DEBUG - Process butonu manuel olarak aktif edildi")
                    
                except Exception as e:
                    print(f"DEBUG - Thread kontrol hatası: {e}")
        else:
            print("DEBUG - display_result zaten çağrılmış, df_result mevcut")

    def display_result(self, df):
        try:
            logging.info("Sonuçlar gösteriliyor...")
            print("DEBUG - display_result çağrıldı")
            print(f"DEBUG - DataFrame boyutu: {df.shape if df is not None else 'None'}")
            print(f"DEBUG - DataFrame tipi: {type(df)}")
            print(f"DEBUG - Current thread: {QThread.currentThread()}")
            print(f"DEBUG - Main thread: {QApplication.instance().thread()}")
            
            # İşlem durumunu güncelle
            if hasattr(self, '_processing'):
                self._processing = False
            
            # Widget'ın hala var olup olmadığını kontrol et
            if not self or not hasattr(self, 'table_widget'):
                logging.error("Widget artık mevcut değil!")
                return

            # Widget'ların silinip silinmediğini kontrol et
            try:
                # Test amaçlı widget'lara erişim dene
                _ = self.table_widget.objectName()
                _ = self.process_btn.objectName()
            except RuntimeError:
                logging.error("Widget'lar silinmiş, işlem iptal ediliyor")
                return
            
            # DataFrame kontrolü
            if df is None:
                print("DEBUG - DataFrame None, display_result iptal ediliyor")
                try:
                    self.process_btn.setEnabled(True)
                except RuntimeError:
                    print("DEBUG - Process button silinmiş")
                return
                
            if df.empty:
                print("DEBUG - DataFrame boş, display_result iptal ediliyor")
                try:
                    self.process_btn.setEnabled(True)
                except RuntimeError:
                    print("DEBUG - Process button silinmiş")
                return
                
            self.df_result = df

            # UI güncellemesi için main thread'de olduğumuzdan emin ol
            QApplication.processEvents()

            # DataFrame'e ID kolonu ekle (eğer yoksa) - bu kolon tabloda görünmeyecek
            if 'harcama_talep_id' not in df.columns:
                df['harcama_talep_id'] = None
            
            # Tabloyu temizle ve yeni verileri yükle (ID kolonu hariç)
            display_columns = [col for col in df.columns if col != 'harcama_talep_id']
            self.table_widget.clear()
            self.table_widget.setColumnCount(len(display_columns))
            self.table_widget.setRowCount(len(df))
            self.table_widget.setHorizontalHeaderLabels(display_columns)

            # Verileri tabloya yükle (ID kolonu hariç)
            for row in range(len(df)):
                display_col = 0
                for col_idx, col_name in enumerate(df.columns):
                    if col_name == 'harcama_talep_id':
                        continue  # ID kolonunu atla
                    
                    try:
                        value = df.iat[row, col_idx]
                        if pd.isna(value):
                            value = ""
                        item = QTableWidgetItem(str(value))
                        # Hücre düzenlenebilir yap (sadece veri sütunları)
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                        self.table_widget.setItem(row, display_col, item)
                    except Exception as cell_error:
                        print(f"DEBUG - Hücre hatası (row={row}, col={display_col}): {cell_error}")
                        item = QTableWidgetItem("")
                        item.setFlags(item.flags() | Qt.ItemIsEditable)
                    self.table_widget.setItem(row, display_col, item)
                    display_col += 1

            # Hücre değişikliklerini dinle (otomatik kayıt için)
            # Önce mevcut bağlantıyı kaldır (tekrar bağlanmamak için)
            try:
                self.table_widget.itemChanged.disconnect()
            except:
                pass
            self.table_widget.itemChanged.connect(self.on_cell_changed)

            # Tabloyu boyutlandır
            self.table_widget.resizeColumnsToContents()
            
            # UI'yi güncelle
            try:
                self.process_btn.setEnabled(True)
                self.save_db_btn.show()
                self.save_output_btn.show()
                self.manual_add_btn.show()
                if hasattr(self, 'progress_bar'):
                    self.progress_bar.setVisible(False)
            except RuntimeError as ui_error:
                print(f"DEBUG - UI güncelleme hatası: {ui_error}")

            # Otomatik kayıt kaldırıldı - kullanıcı "Veritabanına Kaydet" butonuna basacak

            logging.info("Sonuçlar başarıyla gösterildi")
            print("DEBUG - display_result başarıyla tamamlandı")
            
            # Başarı mesajını göster
            try:
                QMessageBox.information(self, "Başarılı", "İşlem tamamlandı ve veriler otomatik olarak veritabanına kaydedildi.")
            except Exception as msg_error:
                print(f"DEBUG - Mesaj gösterme hatası: {msg_error}")

        except Exception as e:
            error_msg = f"Sonuç gösterme hatası: {str(e)}"
            logging.error(error_msg)
            print(f"ERROR - display_result hatası: {str(e)}")
            import traceback
            print(f"ERROR - display_result traceback: {traceback.format_exc()}")
            
            # UI'yi güvenli şekilde güncelle
            try:
                self.process_btn.setEnabled(True)
                if hasattr(self, 'progress_bar'):
                    self.progress_bar.setVisible(False)
            except Exception as ui_error:
                print(f"DEBUG - UI güncelleme hatası: {ui_error}")
            
            # Hata mesajını göster
            try:
                QMessageBox.critical(self, "Hata", f"Sonuç gösterilirken hata: {str(e)}")
            except Exception as msg_error:
                print(f"DEBUG - Hata mesajı gösterme hatası: {msg_error}")
    
    def save_to_database(self):
        """Veritabanına kaydet butonu için"""
        if not hasattr(self, 'df_result') or self.df_result is None:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek veri bulunamadı!")
            return
        
        if not self.user_id:
            QMessageBox.warning(self, "Uyarı", "Kullanıcı ID bulunamadı!")
            return
        
        # Kullanıcıya onay sor
        reply = QMessageBox.question(
            self, 
            "Onay", 
            f"{len(self.df_result)} satır veritabanına kaydedilecek. Devam etmek istiyor musunuz?",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes
        )
        
        if reply != QMessageBox.Yes:
            return
        
        # Butonu geçici olarak devre dışı bırak
        self.save_db_btn.setEnabled(False)
        self.save_db_btn.setText("⏳ Kaydediliyor...")
        
        try:
            # Otomatik kayıt metodunu çağır
            self._auto_save_all_rows()
            
            QMessageBox.information(
                self, 
                "Başarılı", 
                "Veriler başarıyla veritabanına kaydedildi!"
            )
        except Exception as e:
            QMessageBox.critical(
                self, 
                "Hata", 
                f"Veriler kaydedilirken hata oluştu:\n{str(e)}"
            )
            import traceback
            traceback.print_exc()
        finally:
            # Butonu tekrar aktif et
            self.save_db_btn.setEnabled(True)
            self.save_db_btn.setText("💾 Kaydet")
    
    def save_output_file(self):
        if self.df_result is not None:
            try:
                logging.info("Dosya kaydetme işlemi başlatılıyor...")

                # Proje tipini seçtir
                from PyQt5.QtWidgets import QInputDialog
                project_types = ["PMI", "JTI", "TOPPING", "RUSTICA", "İZMİR", "FCV"]
                project_type, ok = QInputDialog.getItem(
                    self, 
                    "Proje Tipi Seçin", 
                    "Lütfen proje tipini seçin:",
                    project_types, 
                    0, 
                    False
                )
                
                if not ok or not project_type:
                    print("Proje tipi seçimi iptal edildi")
                    return

                # Güvenli varsayılan konum
                default_dir = os.path.expanduser("~/Desktop")
                if not os.path.exists(default_dir):
                    default_dir = os.path.expanduser("~")

                default_path = os.path.join(default_dir, f"2025_{project_type}_Proje_Harcama_Talep_Formu.xlsx")

                file_name, _ = QFileDialog.getSaveFileName(
                    self,
                    "Çıktı Dosyasını Kaydet",
                    default_path,
                    "Excel Files (*.xlsx);;All Files (*.*)"
                )

                if file_name:
                    logging.info(f"Dosya kaydediliyor: {file_name}")
                    
                    # Excel dosyasını openpyxl ile oluştur
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils.dataframe import dataframe_to_rows
                    
                    # Yeni workbook oluştur
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Harcama Talep Formu"
                    
                    # Renk tanımları
                    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Olive Green Accent 3 Lighter 80%
                    blue_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")  # Blue Accent 1 Lighter 80%
                    black_font = Font(color="000000", bold=True, size=16)  # Siyah kalın font
                    header_font = Font(color="000000", bold=True)   # Siyah kalın font (başlık için)
                    
                    # Başlık satırı (yeşil arka plan, siyah yazı) - 1. satır
                    title = f"2025 {project_type} Proje Harcama Talep Formu"
                    ws.merge_cells('A1:M1')
                    ws['A1'] = title
                    ws['A1'].font = black_font
                    ws['A1'].fill = green_fill
                    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Başlık satırının yüksekliğini ayarla
                    ws.row_dimensions[1].height = 30
                    
                    # 2. satır boş bırakılıyor
                    
                    # Tablo başlıkları (mavi arka plan, siyah yazı) - 3. satır
                    headers = [
                        "No", "Tarih", "BÖLGE KODU", "KAYNAK TİPİ KODU", "STAGE KODU", 
                        "STAGE-OPERASYON KODU", "Safha", "Harcama Kalemi", "Birim", 
                        "Miktar", "Birim ücret", "Toplam", "Açıklama"
                    ]
                    
                    for col, header in enumerate(headers, 1):
                        cell = ws.cell(row=4, column=col, value=header)
                        cell.font = header_font
                        cell.fill = blue_fill  # Blue Accent 1 Lighter 80% arka plan
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Tablo başlık satırının yüksekliğini ayarla
                    ws.row_dimensions[4].height = 25
                    
                    # Veri satırlarını ekle (4. satırdan başlayarak) - ID kolonu hariç
                    # ID kolonunu çıkar
                    df_for_excel = self.df_result.drop(columns=['harcama_talep_id'], errors='ignore')
                    last_data_row = 4  # Son veri satırını takip etmek için
                    for row_idx, row_data in enumerate(dataframe_to_rows(df_for_excel, index=False, header=False), 5):
                        last_data_row = row_idx
                        for col_idx, value in enumerate(row_data, 1):
                            cell = ws.cell(row=row_idx, column=col_idx, value=value)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            
                            # Birim ücret (11. sütun - K sütunu) için sayısal değer ve format
                            if col_idx == 11:  # Birim ücret sütunu (K sütunu)
                                if value is not None and value != '':
                                    try:
                                        # TL yazısını temizle ve sayıya çevir
                                        if isinstance(value, str):
                                            value = value.replace('TL', '').replace('₺', '').strip()
                                        num_value = float(value)
                                        cell.value = num_value  # Sayısal değer olarak sakla (formül için)
                                        cell.number_format = '#,##0.00" ₺"'  # Görüntüleme formatı
                                    except:
                                        pass
                            elif col_idx == 12:  # Toplam sütunu (L sütunu) - Formül ekle
                                # L sütununa J*K formülü ekle (J=10. sütun, K=11. sütun)
                                from openpyxl.utils import get_column_letter
                                j_col = get_column_letter(10)  # J sütunu (Miktar)
                                k_col = get_column_letter(11)  # K sütunu (Birim ücret)
                                l_col = get_column_letter(12)  # L sütunu (Toplam)
                                
                                # Formül: =J5*K5 (satır numarasına göre)
                                formula = f"={j_col}{row_idx}*{k_col}{row_idx}"
                                cell.value = formula
                                cell.number_format = '#,##0.00" ₺"'  # Sayı formatı
                            elif col_idx == 10:  # Miktar sütunu (J sütunu) - Sayı formatı
                                if value is not None and value != '':
                                    try:
                                        num_value = float(value)
                                        cell.value = num_value
                                        cell.number_format = '#,##0.00'
                                    except:
                                        pass
                    
                    # L sütununun en altına toplam formülü ekle (BOLD)
                    alt_toplam_row = last_data_row + 1
                    l_col_letter = get_column_letter(12)  # L sütunu
                    # L sütunundaki tüm değerlerin toplamı için formül: =SUM(L5:L{last_row})
                    sum_formula = f"=SUM({l_col_letter}5:{l_col_letter}{last_data_row})"
                    alt_toplam_cell = ws.cell(row=alt_toplam_row, column=12, value=sum_formula)
                    alt_toplam_cell.font = Font(bold=True, size=11)
                    alt_toplam_cell.number_format = '#,##0.00" ₺"'  # Sayı formatı
                    alt_toplam_cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    # Sütun genişliklerini otomatik ayarla
                    for col_num in range(1, ws.max_column + 1):
                        max_length = 0
                        # Column letter'ı güvenli şekilde al
                        from openpyxl.utils import get_column_letter
                        column_letter = get_column_letter(col_num)
                        
                        # Sadece veri satırlarını kontrol et (3. satırdan itibaren)
                        for row_num in range(4, ws.max_row + 1):
                            cell = ws.cell(row=row_num, column=col_num)
                            try:
                                if cell.value and len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        # Başlık satırını da kontrol et (2. satır)
                        header_cell = ws.cell(row=3, column=col_num)
                        try:
                            if header_cell.value and len(str(header_cell.value)) > max_length:
                                max_length = len(str(header_cell.value))
                        except:
                            pass
                        
                        adjusted_width = min(max_length + 2, 50)  # Maksimum 50 karakter
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    # Kenarlık ekle
                    thin_border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    
                    # Tüm hücrelere kenarlık ekle
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                        for cell in row:
                            cell.border = thin_border
                    
                    # Dosyayı kaydet
                    wb.save(file_name)
                    wb.close()
                    
                    logging.info("Dosya başarıyla kaydedildi")
                    
                    # Excel'e aktarırken otomatik kayıt kaldırıldı - kullanıcı "Veritabanına Kaydet" butonunu kullanmalı
                    QMessageBox.information(self, "Başarılı", 
                        f"Dosya başarıyla kaydedildi:\n{file_name}\n\nVerileri veritabanına kaydetmek için 'Veritabanına Kaydet' butonunu kullanın.")

            except Exception as e:
                error_msg = f"Dosya kaydetme hatası: {str(e)}"
                logging.error(error_msg)
                print(error_msg)
                import traceback
                print(f"Traceback: {traceback.format_exc()}")
                QMessageBox.critical(self, "Hata", f"Dosya kaydedilemedi:\n{str(e)}")
        else:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek veri bulunamadı.")
    
    def save_to_database(self):
        """Verileri veritabanına kaydet"""
        if self.df_result is None or self.df_result.empty:
            QMessageBox.warning(self, "Uyarı", "Kaydedilecek veri bulunamadı.")
            return
        
        if not self.user_id:
            QMessageBox.warning(self, "Uyarı", "Kullanıcı bilgisi bulunamadı.")
            return
        
        try:
            import requests
            from datetime import datetime
            
            # Her satırı veritabanına kaydet
            success_count = 0
            error_count = 0
            
            for idx, row in self.df_result.iterrows():
                try:
                    # Tarih formatını düzelt
                    tarih = row.get('Tarih', datetime.now().strftime('%Y-%m-%d'))
                    if isinstance(tarih, str):
                        # Eğer tarih string ise parse et
                        try:
                            tarih = pd.to_datetime(tarih).strftime('%Y-%m-%d')
                        except:
                            tarih = datetime.now().strftime('%Y-%m-%d')
                    else:
                        tarih = pd.to_datetime(tarih).strftime('%Y-%m-%d')
                    
                    # Veriyi hazırla
                    data = {
                        'user_id': self.user_id,
                        'tarih': tarih,
                        'bolge_kodu': str(row.get('BÖLGE KODU', '')),
                        'kaynak_tipi_kodu': str(row.get('KAYNAK TİPİ KODU', '')),
                        'stage_kodu': str(row.get('STAGE KODU', '')),
                        'stage_operasyon_kodu': str(row.get('STAGE-OPERASYON KODU', '')),
                        'safha': str(row.get('Safha', '')),
                        'harcama_kalemi': str(row.get('Harcama Kalemi', '')),
                        'birim': str(row.get('Birim', '')),
                        'miktar': float(row.get('Miktar', 0)) if pd.notna(row.get('Miktar')) else 0,
                        'birim_ucret': float(row.get('Birim ücret', 0)) if pd.notna(row.get('Birim ücret')) else 0,
                        'toplam': float(row.get('Toplam', 0)) if pd.notna(row.get('Toplam')) else 0,
                        'aciklama': str(row.get('Açıklama', '')),
                        'is_manuel': 0
                    }
                    
                    # API'ye gönder
                    response = requests.post(
                        f'{get_api_root()}/harcama_talep',
                        json=data,
                        headers=merge_auth_headers(),
                    )
                    
                    if response.status_code == 201:
                        success_count += 1
                    else:
                        error_count += 1
                        print(f"Hata - Satır {idx}: {response.text}")
                        
                except Exception as e:
                    error_count += 1
                    print(f"Hata - Satır {idx}: {str(e)}")
            
            # Sonuç mesajı
            if error_count == 0:
                QMessageBox.information(self, "Başarılı", f"{success_count} kayıt başarıyla veritabanına kaydedildi.")
            else:
                QMessageBox.warning(self, "Kısmen Başarılı", 
                    f"{success_count} kayıt başarıyla kaydedildi.\n{error_count} kayıt kaydedilemedi.")
                
        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Veritabanına kaydetme hatası: {str(e)}")
    
    def on_cell_changed(self, item):
        """Hücre değiştiğinde otomatik kayıt yap"""
        if not hasattr(self, 'df_result') or self.df_result is None:
            return
        
        if not self.user_id:
            return
        
        # Eğer şu anda veritabanına kayıt yapılıyorsa, tekrar kayıt yapma
        if hasattr(self, '_saving_to_db') and self._saving_to_db:
            return
        
        try:
            row = item.row()
            col = item.column()
            
            # Sütun isimlerini al (ID kolonu hariç)
            display_columns = [col for col in self.df_result.columns if col != 'harcama_talep_id']
            if col >= len(display_columns):
                return
            
            column_name = display_columns[col]
            new_value = item.text()
            
            # DataFrame'deki gerçek sütun indeksini bul
            real_col_idx = self.df_result.columns.get_loc(column_name)
            
            # DataFrame'deki değeri güncelle
            if row < len(self.df_result):
                old_value = str(self.df_result.iloc[row, real_col_idx]) if pd.notna(self.df_result.iloc[row, real_col_idx]) else ''
                self.df_result.iloc[row, real_col_idx] = new_value
                
                # Eğer değişiklik varsa veritabanına kaydet
                if old_value != new_value and column_name in ['BÖLGE KODU', 'KAYNAK TİPİ KODU', 'STAGE KODU', 
                                                               'STAGE-OPERASYON KODU', 'Safha', 'Harcama Kalemi', 
                                                               'Birim', 'Miktar', 'Birim ücret', 'Toplam', 'Açıklama']:
                    # Satır verilerini al
                    row_data = self.df_result.iloc[row].to_dict()
                    
                    # Veritabanına kaydet veya güncelle
                    self._save_row_to_database(row, row_data, column_name, old_value, new_value)
                    
        except Exception as e:
            print(f"DEBUG - Hücre değişikliği hatası: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def _auto_save_all_rows(self):
        """Tüm satırları otomatik olarak veritabanına kaydet"""
        print("DEBUG - _auto_save_all_rows çağrıldı")
        
        if not hasattr(self, 'df_result') or self.df_result is None:
            print("DEBUG - df_result yok veya None, kayıt yapılmıyor")
            return
        
        if not self.user_id:
            print(f"DEBUG - user_id yok: {self.user_id}, kayıt yapılmıyor")
            return
        
        print(f"DEBUG - df_result boyutu: {len(self.df_result)} satır")
        print(f"DEBUG - user_id: {self.user_id}")
        
        # Kayıt yapılıyor flag'i
        self._saving_to_db = True
        
        try:
            import requests
            from datetime import datetime
            
            success_count = 0
            error_count = 0
            
            for idx, row in self.df_result.iterrows():
                try:
                    # Tarih formatını düzelt
                    tarih = row.get('Tarih', datetime.now().strftime('%Y-%m-%d'))
                    if isinstance(tarih, str):
                        try:
                            tarih = pd.to_datetime(tarih).strftime('%Y-%m-%d')
                        except:
                            tarih = datetime.now().strftime('%Y-%m-%d')
                    else:
                        tarih = pd.to_datetime(tarih).strftime('%Y-%m-%d')
                    
                    # Bölge kodunu normalize et (nokta varsa kaldır)
                    bolge_kodu = str(row.get('BÖLGE KODU', '')).strip()
                    if bolge_kodu.endswith('.'):
                        bolge_kodu = bolge_kodu[:-1]
                    
                    # Veriyi hazırla
                    data = {
                        'user_id': self.user_id,
                        'tarih': tarih,
                        'bolge_kodu': bolge_kodu,
                        'kaynak_tipi_kodu': str(row.get('KAYNAK TİPİ KODU', '')),
                        'stage_kodu': str(row.get('STAGE KODU', '')),
                        'stage_operasyon_kodu': str(row.get('STAGE-OPERASYON KODU', '')),
                        'safha': str(row.get('Safha', '')),
                        'harcama_kalemi': str(row.get('Harcama Kalemi', '')),
                        'birim': str(row.get('Birim', '')),
                        'miktar': float(row.get('Miktar', 0)) if pd.notna(row.get('Miktar')) else 0,
                        'birim_ucret': float(row.get('Birim ücret', 0)) if pd.notna(row.get('Birim ücret')) else 0,
                        'toplam': float(row.get('Toplam', 0)) if pd.notna(row.get('Toplam')) else 0,
                        'aciklama': str(row.get('Açıklama', '')),
                        'is_manuel': 0  # Otomatik kayıt
                    }
                    
                    # Eğer bu satır daha önce kaydedilmişse (ID varsa), güncelle
                    harcama_talep_id = None
                    if 'harcama_talep_id' in self.df_result.columns:
                        harcama_talep_id = self.df_result.iloc[idx].get('harcama_talep_id')
                        if pd.notna(harcama_talep_id) and harcama_talep_id:
                            harcama_talep_id = int(harcama_talep_id)
                    
                    if harcama_talep_id:
                        # Güncelleme yap
                        response = requests.put(
                            f'{get_api_root()}/harcama_talep/{harcama_talep_id}',
                            json=data,
                            timeout=5,
                            headers=merge_auth_headers(),
                        )
                        if response.status_code == 200:
                            success_count += 1
                        else:
                            error_count += 1
                            print(f"⚠️ Satır {idx} güncellenemedi: {response.text}")
                    else:
                        # Yeni kayıt oluştur
                        response = requests.post(
                            f'{get_api_root()}/harcama_talep',
                            json=data,
                            timeout=5,
                            headers=merge_auth_headers(),
                        )
                        
                        if response.status_code == 201:
                            response_data = response.json()
                            harcama_talep_id = response_data.get('harcama_talep_id')
                            
                            # DataFrame'e ID ekle (güncelleme için)
                            if 'harcama_talep_id' not in self.df_result.columns:
                                self.df_result['harcama_talep_id'] = None
                            self.df_result.iloc[idx, self.df_result.columns.get_loc('harcama_talep_id')] = harcama_talep_id
                            
                            success_count += 1
                        else:
                            error_count += 1
                            print(f"⚠️ Satır {idx} kaydedilemedi: {response.status_code} - {response.text}")
                            print(f"DEBUG - Gönderilen data: {data}")
                        
                except Exception as e:
                    error_count += 1
                    print(f"❌ Satır {idx} kayıt hatası: {str(e)}")
            
            print(f"✅ {success_count} satır otomatik olarak veritabanına kaydedildi.")
            if error_count > 0:
                print(f"⚠️ {error_count} satır kaydedilemedi.")
            
            # Debug: Toplam kayıt sayısını göster
            print(f"DEBUG - Toplam işlenen satır: {len(self.df_result)}")
            print(f"DEBUG - Başarılı kayıt: {success_count}, Hatalı kayıt: {error_count}")
                
        except Exception as e:
            print(f"❌ Otomatik kayıt hatası: {str(e)}")
            import traceback
            traceback.print_exc()
        finally:
            self._saving_to_db = False
    
    def _save_row_to_database(self, row_idx, row_data, changed_column, old_value, new_value):
        """Satırı veritabanına kaydet veya güncelle"""
        try:
            import requests
            from datetime import datetime
            
            # Tarih formatını düzelt
            tarih = row_data.get('Tarih', datetime.now().strftime('%Y-%m-%d'))
            if isinstance(tarih, str):
                try:
                    tarih = pd.to_datetime(tarih).strftime('%Y-%m-%d')
                except:
                    tarih = datetime.now().strftime('%Y-%m-%d')
            else:
                tarih = pd.to_datetime(tarih).strftime('%Y-%m-%d')
            
            # Veriyi hazırla
            data = {
                'user_id': self.user_id,
                'tarih': tarih,
                'bolge_kodu': str(row_data.get('BÖLGE KODU', '')),
                'kaynak_tipi_kodu': str(row_data.get('KAYNAK TİPİ KODU', '')),
                'stage_kodu': str(row_data.get('STAGE KODU', '')),
                'stage_operasyon_kodu': str(row_data.get('STAGE-OPERASYON KODU', '')),
                'safha': str(row_data.get('Safha', '')),
                'harcama_kalemi': str(row_data.get('Harcama Kalemi', '')),
                'birim': str(row_data.get('Birim', '')),
                'miktar': float(row_data.get('Miktar', 0)) if pd.notna(row_data.get('Miktar')) else 0,
                'birim_ucret': float(row_data.get('Birim ücret', 0)) if pd.notna(row_data.get('Birim ücret')) else 0,
                'toplam': float(row_data.get('Toplam', 0)) if pd.notna(row_data.get('Toplam')) else 0,
                'aciklama': str(row_data.get('Açıklama', '')),
                'is_manuel': 1  # Manuel değişiklik
            }
            
            # Eğer bu satır daha önce kaydedilmişse (ID varsa), güncelle
            harcama_talep_id = None
            if 'harcama_talep_id' in self.df_result.columns:
                harcama_talep_id = self.df_result.iloc[row_idx].get('harcama_talep_id')
                if pd.notna(harcama_talep_id) and harcama_talep_id:
                    harcama_talep_id = int(harcama_talep_id)
            
            if harcama_talep_id:
                # Mevcut kaydı güncelle
                response = requests.put(
                    f'{get_api_root()}/harcama_talep/{harcama_talep_id}',
                    json=data,
                    timeout=5,
                    headers=merge_auth_headers(),
                )
                if response.status_code == 200:
                    print(f"✅ Satır {row_idx} (ID: {harcama_talep_id}) otomatik olarak güncellendi.")
                else:
                    print(f"⚠️ Satır {row_idx} güncellenemedi: {response.text}")
            else:
                # Yeni kayıt ekle
                response = requests.post(
                    f'{get_api_root()}/harcama_talep',
                    json=data,
                    timeout=5,
                    headers=merge_auth_headers(),
                )
                if response.status_code == 201:
                    response_data = response.json()
                    new_id = response_data.get('harcama_talep_id')
                    
                    # DataFrame'e ID ekle
                    if 'harcama_talep_id' not in self.df_result.columns:
                        self.df_result['harcama_talep_id'] = None
                    self.df_result.iloc[row_idx, self.df_result.columns.get_loc('harcama_talep_id')] = new_id
                    
                    print(f"✅ Satır {row_idx} otomatik olarak veritabanına kaydedildi (ID: {new_id}).")
                else:
                    print(f"⚠️ Satır {row_idx} kaydedilemedi: {response.text}")
                
        except Exception as e:
            print(f"❌ Otomatik kayıt/güncelleme hatası: {str(e)}")
            import traceback
            traceback.print_exc()
    
    def show_manual_add_dialog(self):
        """Manuel ekleme dialog'unu göster"""
        dialog = ManualAddDialog(self.api_client, self.user_id, self)
        if dialog.exec_() == QDialog.Accepted:
            # Dialog'dan eklenen veriyi al ve tabloya ekle
            new_data = dialog.get_data()
            if new_data:
                # DataFrame'e ekle
                new_row = pd.DataFrame([new_data])
                if self.df_result is None:
                    self.df_result = new_row
                else:
                    self.df_result = pd.concat([self.df_result, new_row], ignore_index=True)
                
                # Tabloyu güncelle
                self.display_result(self.df_result)
                QMessageBox.information(self, "Başarılı", "Manuel kayıt eklendi.")


class ManualAddDialog(QDialog):
    """Manuel ekleme dialog'u"""
    def __init__(self, api_client, user_id, parent=None):
        super().__init__(parent)
        self.api_client = api_client
        self.user_id = user_id
        self.setWindowTitle("Manuel Harcama Talep Ekle")
        self.setMinimumWidth(600)
        self.init_ui()
    
    def init_ui(self):
        layout = QVBoxLayout(self)
        
        form_layout = QFormLayout()
        
        # Tarih
        self.tarih_edit = QDateEdit()
        self.tarih_edit.setDate(QDate.currentDate())
        self.tarih_edit.setCalendarPopup(True)
        form_layout.addRow("Tarih:", self.tarih_edit)
        
        # Bölge Kodu
        self.bolge_kodu_combo = QComboBox()
        # Bölge kodlarını API'den al
        try:
            import requests
            response = requests.get(
                f"{get_api_root()}/bolge_kodlari?user_id={self.user_id}",
                headers=merge_auth_headers(),
            )
            if response.status_code == 200:
                bolge_kodlari = response.json()
                for kod, ad in bolge_kodlari.items():
                    self.bolge_kodu_combo.addItem(f"{ad} ({kod})", kod)
        except:
            pass
        form_layout.addRow("Bölge Kodu:", self.bolge_kodu_combo)
        
        # Kaynak Tipi Kodu
        self.kaynak_tipi_combo = QComboBox()
        try:
            import requests
            response = requests.get(
                f"{get_api_root()}/kaynak_tipleri",
                headers=merge_auth_headers(),
            )
            if response.status_code == 200:
                kaynak_tipleri = response.json()
                for kod, ad in kaynak_tipleri.items():
                    self.kaynak_tipi_combo.addItem(f"{ad} ({kod})", kod)
        except:
            pass
        form_layout.addRow("Kaynak Tipi Kodu:", self.kaynak_tipi_combo)
        
        # Stage Kodu
        self.stage_combo = QComboBox()
        try:
            import requests
            response = requests.get(
                f"{get_api_root()}/stages",
                headers=merge_auth_headers(),
            )
            if response.status_code == 200:
                stages = response.json()
                for kod, ad in stages.items():
                    self.stage_combo.addItem(f"{ad} ({kod})", kod)
        except:
            pass
        form_layout.addRow("Stage Kodu:", self.stage_combo)
        
        # Stage-Operasyon Kodu
        self.stage_operasyon_edit = QLineEdit()
        form_layout.addRow("Stage-Operasyon Kodu:", self.stage_operasyon_edit)
        
        # Safha
        self.safha_edit = QLineEdit()
        form_layout.addRow("Safha:", self.safha_edit)
        
        # Harcama Kalemi
        self.harcama_kalemi_edit = QLineEdit()
        form_layout.addRow("Harcama Kalemi:", self.harcama_kalemi_edit)
        
        # Birim
        self.birim_edit = QLineEdit()
        form_layout.addRow("Birim:", self.birim_edit)
        
        # Miktar
        self.miktar_spin = QDoubleSpinBox()
        self.miktar_spin.setMaximum(999999999)
        form_layout.addRow("Miktar:", self.miktar_spin)
        
        # Birim Ücret
        self.birim_ucret_spin = QDoubleSpinBox()
        self.birim_ucret_spin.setMaximum(999999999)
        form_layout.addRow("Birim Ücret:", self.birim_ucret_spin)
        
        # Toplam
        self.toplam_spin = QDoubleSpinBox()
        self.toplam_spin.setMaximum(999999999)
        form_layout.addRow("Toplam:", self.toplam_spin)
        
        # Açıklama
        self.aciklama_edit = QLineEdit()
        form_layout.addRow("Açıklama:", self.aciklama_edit)
        
        layout.addLayout(form_layout)
        
        # Butonlar
        buttons = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def get_data(self):
        """Dialog'dan veriyi al"""
        from datetime import datetime
        
        # Toplam hesapla
        miktar = self.miktar_spin.value()
        birim_ucret = self.birim_ucret_spin.value()
        toplam = miktar * birim_ucret
        self.toplam_spin.setValue(toplam)
        
        return {
            'Tarih': self.tarih_edit.date().toString('yyyy-MM-dd'),
            'BÖLGE KODU': self.bolge_kodu_combo.currentData() or '',
            'KAYNAK TİPİ KODU': self.kaynak_tipi_combo.currentData() or '',
            'STAGE KODU': self.stage_combo.currentData() or '',
            'STAGE-OPERASYON KODU': self.stage_operasyon_edit.text(),
            'Safha': self.safha_edit.text(),
            'Harcama Kalemi': self.harcama_kalemi_edit.text(),
            'Birim': self.birim_edit.text(),
            'Miktar': miktar,
            'Birim ücret': birim_ucret,
            'Toplam': toplam,
            'Açıklama': self.aciklama_edit.text()
        }

    def show_error(self, error_message):
        error_msg = f"Thread hatası: {error_message}"
        logging.error(error_msg)
        print(f"DEBUG - show_error çağrıldı: {error_message}")
        QMessageBox.critical(self, "Hata", f"İşlem sırasında bir hata oluştu:\n{error_message}")
        self.process_btn.setEnabled(True)
        self.save_output_btn.hide()