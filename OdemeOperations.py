import pandas as pd
import numpy as np
from datetime import datetime
import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QPushButton,
                             QLabel, QFileDialog, QMessageBox, QTextEdit,
                             QFrame, QGridLayout, QProgressBar, QDateEdit,
                             QLineEdit, QComboBox, QTableWidget, QTableWidgetItem,
                             QHeaderView, QGroupBox, QSpacerItem, QSizePolicy)
from PyQt5.QtCore import Qt, QThread, pyqtSignal, QDate
from PyQt5.QtGui import QFont, QIcon


class ExcelProcessorThread(QThread):
    progress_updated = pyqtSignal(int)
    log_updated = pyqtSignal(str)
    finished_processing = pyqtSignal(list, float)
    error_occurred = pyqtSignal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path
        self.dayibasi_data = []

    def run(self):
        try:
            self.log_updated.emit("🔍 Excel dosyası kontrol ediliyor...")

            # Dosya varlık kontrolü
            if not os.path.exists(self.file_path):
                self.error_occurred.emit(f"Dosya bulunamadı: {self.file_path}")
                return

            self.log_updated.emit("📖 Excel dosyası okunuyor...")
            self.progress_updated.emit(10)

            # Excel dosyasını oku - daha güvenli bir şekilde
            try:
                # Engine parametresi eklendi - openpyxl daha güvenli
                excel_file = pd.ExcelFile(self.file_path, engine='openpyxl')
                self.log_updated.emit(f"✅ Excel dosyası başarıyla açıldı")
                self.log_updated.emit(f"📊 Toplam sayfa sayısı: {len(excel_file.sheet_names)}")
            except Exception as e:
                self.error_occurred.emit(f"Excel dosyası okunamadı: {str(e)}")
                return

            self.progress_updated.emit(20)

            # Sayfa sayfa işle
            processed_sheets = 0
            for i, sheet_name in enumerate(excel_file.sheet_names):
                try:
                    self.log_updated.emit(f"📋 Sayfa işleniyor ({i + 1}/{len(excel_file.sheet_names)}): {sheet_name}")

                    # DataFrame oku - daha güvenli parametreler
                    df = pd.read_excel(
                        self.file_path,
                        sheet_name=sheet_name,
                        header=None,
                        engine='openpyxl',
                        na_filter=False  # NaN değerleri boş string olarak oku
                    )

                    self.log_updated.emit(f"   📏 Sayfa boyutu: {df.shape[0]} satır x {df.shape[1]} sütun")

                    # Boş sayfaları atla
                    if df.empty or df.shape[0] == 0:
                        self.log_updated.emit(f"   ⚠️ Boş sayfa atlanıyor")
                        continue

                    # Veri çıkar
                    initial_count = len(self.dayibasi_data)
                    self.extract_dayibasi_info(df)
                    found_count = len(self.dayibasi_data) - initial_count

                    if found_count > 0:
                        self.log_updated.emit(f"   ✅ {found_count} dayıbaşı bulundu")
                    else:
                        self.log_updated.emit(f"   ⚠️ Bu sayfada dayıbaşı bulunamadı")

                    processed_sheets += 1
                    progress = 20 + (processed_sheets / len(excel_file.sheet_names)) * 60
                    self.progress_updated.emit(int(progress))

                except Exception as e:
                    self.log_updated.emit(f"   ❌ Sayfa işlenirken hata: {str(e)}")
                    continue

            # Excel dosyasını kapat
            try:
                excel_file.close()
            except:
                pass

            self.progress_updated.emit(80)

            # Sonuçları değerlendir
            if not self.dayibasi_data:
                self.log_updated.emit("⚠️ Hiçbir geçerli dayıbaşı verisi bulunamadı!")
                self.log_updated.emit("💡 Kontrol edilecek noktalar:")
                self.log_updated.emit("   - 'Dayıbaşı Ad-Soyad:' metni var mı?")
                self.log_updated.emit("   - İsim alanı dolu mu?")
                self.log_updated.emit("   - 'Ücret Toplam' satırında geçerli tutar var mı?")

            # Toplam tutarı hesapla
            total_amount = sum([data.get('amount', 0) for data in self.dayibasi_data])

            self.log_updated.emit("=" * 50)
            self.log_updated.emit(f"🎉 İşlem tamamlandı!")
            self.log_updated.emit(f"📊 Toplam işlenen sayfa: {processed_sheets}")
            self.log_updated.emit(f"👥 Bulunan geçerli dayıbaşı: {len(self.dayibasi_data)}")
            self.log_updated.emit(f"💰 Toplam ödeme tutarı: ₺{total_amount:,.2f}")
            self.log_updated.emit("=" * 50)

            self.progress_updated.emit(100)
            self.finished_processing.emit(self.dayibasi_data, total_amount)

        except Exception as e:
            import traceback
            error_detail = traceback.format_exc()
            self.log_updated.emit(f"💥 Beklenmeyen hata:")
            self.log_updated.emit(f"   Hata: {str(e)}")
            self.log_updated.emit(f"   Detay: {error_detail}")
            self.error_occurred.emit(f"Thread çalıştırılırken kritik hata: {str(e)}")
        finally:
            # Thread temizlik işlemleri
            self.log_updated.emit("🔄 Thread temizleniyor...")

    def extract_dayibasi_info(self, df):
        """DataFrame'den dayıbaşı bilgilerini çıkarır - Eksik verileri atlar"""
        current_dayibasi = {}

        for idx, row in df.iterrows():
            try:
                # Güvenli değer çıkarma
                row_values = []
                for cell in row:
                    if cell is not None and str(cell).strip() and str(cell).strip().lower() != 'nan':
                        row_values.append(str(cell).strip())

                if not row_values:  # Boş satırları atla
                    continue

                row_str = ' '.join(row_values)

                # Dayıbaşı Ad-Soyad kontrolü
                if 'Dayıbaşı Ad-Soyad:' in row_str:
                    # Önceki dayıbaşı geçerli ise kaydet
                    if self.is_valid_dayibasi(current_dayibasi):
                        self.dayibasi_data.append(current_dayibasi.copy())
                        self.log_updated.emit(
                            f"✅ Geçerli dayıbaşı kaydedildi: {current_dayibasi.get('name', 'İsim yok')}")

                    # Yeni dayıbaşı başlat
                    current_dayibasi = {}

                    # İsmi çıkar - Türkçe karakterleri de destekle
                    name_match = re.search(r'Dayıbaşı Ad-Soyad:\s*([A-ZÜĞŞIİÖÇa-züğşıiöç\s]+)', row_str, re.IGNORECASE)
                    if name_match:
                        name = name_match.group(1).strip()
                        if name and name.lower() != 'nan' and len(name) > 1:  # Geçerli isim kontrolü
                            current_dayibasi['name'] = name

                # IBAN kontrolü
                elif 'Dayıbaşı IBAN No:' in row_str and 'name' in current_dayibasi:
                    # IBAN formatını ara
                    iban_match = re.search(r'TR\d{2}[\s\d]{20,30}', row_str)
                    if iban_match:
                        iban = re.sub(r'\s+', '', iban_match.group(0))  # Boşlukları kaldır
                        if len(iban) == 26:  # Türkiye IBAN uzunluğu
                            current_dayibasi['iban'] = iban

                    # Banka adını da çıkar
                    bank_patterns = [
                        r'(FİNANSBANK|AKBANK|İNG\s*BANK|FİNANS|QNB|GARANTI|HALKBANK|ZIRAAT|İŞ\s*BANK|TÜRKIYE\s*İŞ\s*BANKASI)',
                        r'(VAKIFBANK|DENİZBANK|TEB|HSBC|ING|YAPI\s*KREDİ)',
                    ]

                    for pattern in bank_patterns:
                        bank_match = re.search(pattern, row_str, re.IGNORECASE)
                        if bank_match:
                            current_dayibasi['bank'] = bank_match.group(1).strip()
                            break

                # Telefon kontrolü
                elif 'Dayıbaşı TEL No:' in row_str and 'name' in current_dayibasi:
                    # Telefon numarası formatları
                    tel_patterns = [
                        r'0?(\d{3}\s*\d{3}\s*\d{2}\s*\d{2})',
                        r'0?(\d{3}\s*\d{3}\s*\d{4})',
                        r'(\d{4}\s*\d{3}\s*\d{2}\s*\d{2})',
                    ]

                    for pattern in tel_patterns:
                        tel_match = re.search(pattern, row_str)
                        if tel_match:
                            phone = tel_match.group(1).strip()
                            if len(phone.replace(' ', '')) >= 10:
                                current_dayibasi['phone'] = phone
                                break

                # Ücret Toplam kontrolü (ödenecek tutar)
                elif any(keyword in row_str for keyword in
                         ['Ücret Toplam', 'ÜCRET TOPLAM', 'Genel Toplam']) and 'name' in current_dayibasi:
                    # Sayıları bul
                    number_patterns = [
                        r'(\d{1,3}(?:\.\d{3})*(?:,\d{2})?)',  # 195.600,00 formatı
                        r'(\d{3,}(?:\.\d{3})*)',  # 195600 veya 195.600 formatı
                        r'(\d+(?:,\d{3})*(?:\.\d{2})?)',  # 195,600.00 formatı
                    ]

                    for pattern in number_patterns:
                        numbers = re.findall(pattern, row_str)
                        if numbers:
                            for num_str in reversed(numbers):  # Son (en büyük) sayıyı al
                                try:
                                    # Farklı formatları normalize et
                                    if ',' in num_str and '.' in num_str:
                                        # Amerikan formatı: 1,234.56
                                        clean_num = num_str.replace(',', '')
                                        amount = float(clean_num)
                                    elif '.' in num_str and len(num_str.split('.')[-1]) == 2:
                                        # Decimal nokta: 1234.56
                                        amount = float(num_str)
                                    elif '.' in num_str:
                                        # Binlik ayırıcı nokta: 1.234 -> 1234
                                        clean_num = num_str.replace('.', '')
                                        amount = float(clean_num)
                                    elif ',' in num_str:
                                        # Türk formatı: 1234,56 -> 1234.56
                                        clean_num = num_str.replace(',', '.')
                                        amount = float(clean_num)
                                    else:
                                        amount = float(num_str)

                                    # Geçerli tutar kontrolü
                                    if amount >= 100:
                                        current_dayibasi['amount'] = amount
                                        break
                                except (ValueError, AttributeError):
                                    continue
                            if 'amount' in current_dayibasi:
                                break

            except Exception as e:
                # Satır işleme hatası - devam et
                self.log_updated.emit(f"⚠️ Satır {idx} işlenirken hata (atlandı): {str(e)[:50]}...")
                continue

        # Son dayıbaşı bilgisini de kontrol et ve geçerli ise kaydet
        if self.is_valid_dayibasi(current_dayibasi):
            self.dayibasi_data.append(current_dayibasi)
            self.log_updated.emit(f"✅ Son dayıbaşı kaydedildi: {current_dayibasi.get('name', 'İsim yok')}")

    def is_valid_dayibasi(self, dayibasi):
        """Dayıbaşı verilerinin geçerli olup olmadığını kontrol eder"""
        if not isinstance(dayibasi, dict):
            return False

        # En az isim ve tutar olmalı
        required_fields = ['name', 'amount']
        for field in required_fields:
            if field not in dayibasi or not dayibasi[field]:
                return False

        # İsim kontrolü
        name = dayibasi.get('name', '').strip()
        if len(name) < 2 or name.lower() in ['nan', 'none', '']:
            return False

        # Tutar kontrolü
        amount = dayibasi.get('amount', 0)
        try:
            amount = float(amount)
            if amount < 100:  # 100 TL'den küçük tutarları geçersiz say
                return False
        except:
            return False

        # IBAN varsa kontrol et
        iban = dayibasi.get('iban', '')
        if iban and (len(iban) != 26 or not iban.startswith('TR')):
            return False

        return True


class PaymentTab(QWidget):
    def __init__(self, api_client=None):
        super().__init__()
        self.api_client = api_client
        self.dayibasi_data = []
        self.total_amount = 0
        self.processor_thread = None  # Thread referansını başlangıçta None yap
        self.file_path = None  # Dosya yolunu başlangıçta None yap

        self.init_ui()

    def init_ui(self):
        """UI bileşenlerini oluştur"""
        main_layout = QVBoxLayout(self)
        main_layout.setSpacing(20)

        # Başlık
        title = QLabel("📋 ÖDEME İŞLEMLERİ")
        title.setAlignment(Qt.AlignCenter)
        title.setStyleSheet("""
            QLabel {
                font-size: 28px;
                font-weight: bold;
                color: #6366f1;
                margin: 20px;
                padding: 15px;
                background: rgba(99, 102, 241, 0.1);
                border-radius: 15px;
                border: 2px solid rgba(99, 102, 241, 0.3);
            }
        """)
        main_layout.addWidget(title)

        # Ana içerik için scroll area
        content_layout = QHBoxLayout()

        # Sol panel - Dosya işlemleri
        left_panel = self.create_left_panel()
        content_layout.addWidget(left_panel, 1)

        # Sağ panel - Önizleme ve rapor
        right_panel = self.create_right_panel()
        content_layout.addWidget(right_panel, 2)

        main_layout.addLayout(content_layout)

    def create_left_panel(self):
        """Sol panel - Dosya yükleme ve ayarlar"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(255, 255, 255, 0.1);
                border-radius: 15px;
                padding: 20px;
            }
        """)

        layout = QVBoxLayout(panel)

        # Dosya seçme bölümü
        file_group = QGroupBox("📁 Excel Dosyası Seçimi")
        file_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #ffffff;
                border: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 10px;
                margin: 10px 0;
                padding-top: 10px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 10px 0 10px;
            }
        """)
        file_layout = QVBoxLayout(file_group)

        # Dosya yolu gösterimi
        self.file_path_label = QLabel("Dosya seçilmedi")
        self.file_path_label.setStyleSheet("""
            QLabel {
                background: rgba(255, 255, 255, 0.1);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 8px;
                padding: 10px;
                color: #8892b0;
                min-height: 20px;
            }
        """)
        file_layout.addWidget(self.file_path_label)

        # Dosya seç butonu
        self.select_file_btn = QPushButton("📂 Excel Dosyası Seç")
        self.select_file_btn.clicked.connect(self.select_excel_file)
        self.select_file_btn.setStyleSheet(self.get_button_style("#6366f1"))
        file_layout.addWidget(self.select_file_btn)

        layout.addWidget(file_group)

        # Tarih ayarları
        date_group = QGroupBox("📅 Rapor Tarih Aralığı")
        date_group.setStyleSheet(file_group.styleSheet())
        date_layout = QGridLayout(date_group)

        # Başlangıç tarihi
        date_layout.addWidget(QLabel("Başlangıç:"), 0, 0)
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate().addDays(-7))
        self.start_date.setStyleSheet(self.get_input_style())
        date_layout.addWidget(self.start_date, 0, 1)

        # Bitiş tarihi
        date_layout.addWidget(QLabel("Bitiş:"), 1, 0)
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate())
        self.end_date.setStyleSheet(self.get_input_style())
        date_layout.addWidget(self.end_date, 1, 1)

        layout.addWidget(date_group)

        # Bölge seçimi
        region_group = QGroupBox("🌍 Bölge Bilgisi")
        region_group.setStyleSheet(file_group.styleSheet())
        region_layout = QVBoxLayout(region_group)

        self.region_combo = QComboBox()
        self.region_combo.addItems(["EGE", "MARMARA", "İÇ ANADOLU", "AKDENİZ"])
        self.region_combo.setStyleSheet(self.get_input_style())
        region_layout.addWidget(self.region_combo)

        layout.addWidget(region_group)

        # İşlem butonları
        button_group = QGroupBox("⚙️ İşlemler")
        button_group.setStyleSheet(file_group.styleSheet())
        button_layout = QVBoxLayout(button_group)

        # Dosyayı işle butonu
        self.process_btn = QPushButton("🔄 Dosyayı İşle")
        self.process_btn.clicked.connect(self.process_excel_file)
        self.process_btn.setEnabled(False)
        self.process_btn.setStyleSheet(self.get_button_style("#8b5cf6"))
        button_layout.addWidget(self.process_btn)

        # Rapor oluştur butonu
        self.create_report_btn = QPushButton("📊 Ödeme Raporu Oluştur")
        self.create_report_btn.clicked.connect(self.create_payment_report)
        self.create_report_btn.setEnabled(False)
        self.create_report_btn.setStyleSheet(self.get_button_style("#10b981"))
        button_layout.addWidget(self.create_report_btn)

        layout.addWidget(button_group)

        # İlerleme çubuğu
        self.progress_bar = QProgressBar()
        self.progress_bar.setVisible(False)
        self.progress_bar.setStyleSheet("""
            QProgressBar {
                border: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 10px;
                text-align: center;
                color: white;
                font-weight: bold;
                background: rgba(255, 255, 255, 0.1);
            }
            QProgressBar::chunk {
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 #6366f1, stop:1 #8b5cf6);
                border-radius: 8px;
            }
        """)
        layout.addWidget(self.progress_bar)

        # Spacer
        layout.addSpacerItem(QSpacerItem(20, 40, QSizePolicy.Minimum, QSizePolicy.Expanding))

        return panel

    def create_right_panel(self):
        """Sağ panel - Veri önizleme ve log"""
        panel = QFrame()
        panel.setStyleSheet("""
            QFrame {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(255, 255, 255, 0.1);
                border-radius: 15px;
                padding: 20px;
            }
        """)

        layout = QVBoxLayout(panel)

        # Özet bilgiler
        summary_group = QGroupBox("📈 Özet Bilgiler")
        summary_group.setStyleSheet("""
            QGroupBox {
                font-size: 16px;
                font-weight: bold;
                color: #ffffff;
                border: 2px solid rgba(255, 255, 255, 0.2);
                border-radius: 10px;
                margin: 10px 0;
                padding-top: 10px;
            }
        """)
        summary_layout = QGridLayout(summary_group)

        # Özet etiketleri
        self.total_dayibasi_label = QLabel("Toplam Dayıbaşı: 0")
        self.total_amount_label = QLabel("Toplam Tutar: ₺0,00")

        for label in [self.total_dayibasi_label, self.total_amount_label]:
            label.setStyleSheet("""
                QLabel {
                    background: rgba(255, 255, 255, 0.1);
                    border-radius: 8px;
                    padding: 10px;
                    color: #ffffff;
                    font-size: 14px;
                    font-weight: bold;
                }
            """)

        summary_layout.addWidget(self.total_dayibasi_label, 0, 0)
        summary_layout.addWidget(self.total_amount_label, 0, 1)

        layout.addWidget(summary_group)

        # Veri tablosu
        table_group = QGroupBox("📋 Dayıbaşı Verileri")
        table_group.setStyleSheet(summary_group.styleSheet())
        table_layout = QVBoxLayout(table_group)

        self.data_table = QTableWidget()
        self.data_table.setColumnCount(5)
        self.data_table.setHorizontalHeaderLabels(["Ad Soyad", "IBAN", "Banka", "Telefon", "Ödeme Tutarı"])

        # Tablo stilini ayarla
        self.data_table.setStyleSheet("""
            QTableWidget {
                background: rgba(255, 255, 255, 0.05);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 10px;
                gridline-color: rgba(255, 255, 255, 0.1);
                color: white;
                selection-background-color: rgba(99, 102, 241, 0.3);
            }
            QTableWidget::item {
                padding: 8px;
                border-bottom: 1px solid rgba(255, 255, 255, 0.1);
            }
            QHeaderView::section {
                background: rgba(99, 102, 241, 0.2);
                color: white;
                padding: 10px;
                border: 1px solid rgba(255, 255, 255, 0.2);
                font-weight: bold;
            }
        """)

        # Sütun genişliklerini ayarla
        header = self.data_table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.Stretch)  # Ad Soyad
        header.setSectionResizeMode(1, QHeaderView.ResizeToContents)  # IBAN
        header.setSectionResizeMode(2, QHeaderView.ResizeToContents)  # Banka
        header.setSectionResizeMode(3, QHeaderView.ResizeToContents)  # Telefon
        header.setSectionResizeMode(4, QHeaderView.ResizeToContents)  # Tutar

        table_layout.addWidget(self.data_table)
        layout.addWidget(table_group)

        # Log alanı
        log_group = QGroupBox("📝 İşlem Logları")
        log_group.setStyleSheet(summary_group.styleSheet())
        log_layout = QVBoxLayout(log_group)

        self.log_text = QTextEdit()
        self.log_text.setMaximumHeight(150)
        self.log_text.setStyleSheet("""
            QTextEdit {
                background: rgba(0, 0, 0, 0.3);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 8px;
                color: #8892b0;
                font-family: 'Courier New', monospace;
                font-size: 12px;
                padding: 10px;
            }
        """)
        log_layout.addWidget(self.log_text)

        layout.addWidget(log_group)

        return panel

    def get_button_style(self, color):
        """Buton stili döndür"""
        return f"""
            QPushButton {{
                background: {color};
                color: white;
                font-size: 14px;
                font-weight: bold;
                border-radius: 10px;
                padding: 12px 20px;
                min-height: 20px;
                border: none;
            }}
            QPushButton:hover {{
                background: qlineargradient(x1:0, y1:0, x2:1, y2:0,
                    stop:0 {color}, stop:1 #a855f7);
                transform: translateY(-2px);
            }}
            QPushButton:pressed {{
                background: {color};
                transform: translateY(0px);
            }}
            QPushButton:disabled {{
                background: rgba(255, 255, 255, 0.1);
                color: rgba(255, 255, 255, 0.3);
            }}
        """

    def get_input_style(self):
        """Input stili döndür"""
        return """
            QDateEdit, QComboBox {
                background: rgba(255, 255, 255, 0.1);
                border: 1px solid rgba(255, 255, 255, 0.2);
                border-radius: 8px;
                padding: 8px;
                color: white;
                font-size: 14px;
                min-height: 20px;
            }
            QDateEdit:focus, QComboBox:focus {
                border: 2px solid #6366f1;
            }
            QComboBox::drop-down {
                border: none;
                background: rgba(255, 255, 255, 0.1);
                border-radius: 4px;
            }
            QComboBox::down-arrow {
                image: none;
                border: 2px solid white;
                width: 6px;
                height: 6px;
                border-top: none;
                border-right: none;
                transform: rotate(-45deg);
                margin-right: 8px;
            }
        """

    def select_excel_file(self):
        """Excel dosyası seç - Güvenli versiyon"""
        try:
            file_path, _ = QFileDialog.getOpenFileName(
                self,
                "Excel Dosyası Seçin",
                "",
                "Excel Dosyaları (*.xlsx *.xls);;Tüm Dosyalar (*)"
            )

            if not file_path:
                return

            # Dosya varlığı kontrolü
            if not os.path.exists(file_path):
                QMessageBox.critical(self, "Hata", "Seçilen dosya bulunamadı!")
                return

            # Dosya boyutu kontrolü (100MB limit)
            try:
                file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
                if file_size > 100:
                    QMessageBox.warning(self, "Uyarı",
                                        f"Dosya çok büyük ({file_size:.1f} MB).\n"
                                        "Daha küçük bir dosya seçin.")
                    return
            except OSError:
                QMessageBox.critical(self, "Hata", "Dosya boyutu okunamadı!")
                return

            # Dosya formatı kontrolü
            file_ext = os.path.splitext(file_path)[1].lower()
            if file_ext not in ['.xlsx', '.xls']:
                QMessageBox.warning(self, "Uyarı",
                                    "Desteklenen formatlar: .xlsx, .xls")
                return

            # Excel dosyasını test et
            try:
                # Basit okuma testi
                test_df = pd.read_excel(file_path, nrows=5, engine='openpyxl')
                if test_df.empty:
                    QMessageBox.warning(self, "Uyarı", "Excel dosyası boş görünüyor.")
                    return
            except Exception as e:
                QMessageBox.critical(self, "Excel Hatası",
                                     f"Excel dosyası okunamadı:\n{str(e)}\n\n"
                                     "Dosyanın açık olmadığından emin olun.")
                return

            # Başarılı - dosya bilgilerini güncelle
            self.file_path = file_path
            self.file_path_label.setText(f"✅ {os.path.basename(file_path)}")
            self.process_btn.setEnabled(True)

            # Log temizle ve yeni bilgileri ekle
            self.log_text.clear()
            self.log_text.append(f"📁 Dosya seçildi: {os.path.basename(file_path)}")
            self.log_text.append(f"📏 Boyut: {file_size:.1f} MB")
            self.log_text.append("✅ Dosya geçerli - İşleme hazır")

        except Exception as e:
            QMessageBox.critical(self, "Beklenmeyen Hata",
                                 f"Dosya seçimi sırasında hata:\n{str(e)}")
            self.log_text.append(f"❌ Dosya seçim hatası: {str(e)}")

    def process_excel_file(self):
        """Excel dosyasını işle - Güvenli versiyon"""
        try:
            # Ön kontroller
            if not hasattr(self, 'file_path') or not self.file_path:
                QMessageBox.warning(self, "Uyarı", "Lütfen önce bir Excel dosyası seçin.")
                return

            if not os.path.exists(self.file_path):
                QMessageBox.critical(self, "Hata", "Seçilen dosya artık mevcut değil!")
                return

            # Önceki thread'i temizle
            self._cleanup_thread()

            # UI hazırlık
            self.log_text.clear()
            self.log_text.append("🚀 Excel dosyası işleme başlatılıyor...")
            self.progress_bar.setVisible(True)
            self.progress_bar.setValue(0)
            self.process_btn.setEnabled(False)
            self.create_report_btn.setEnabled(False)

            # Mevcut verileri temizle
            self.dayibasi_data.clear()
            self.total_amount = 0
            self.data_table.setRowCount(0)
            self.total_dayibasi_label.setText("Toplam Dayıbaşı: 0")
            self.total_amount_label.setText("Toplam Tutar: ₺0,00")

            # Thread oluştur ve başlat
            self.processor_thread = ExcelProcessorThread(self.file_path)

            # Signal bağlantıları
            self.processor_thread.progress_updated.connect(self.update_progress)
            self.processor_thread.log_updated.connect(self.add_log_message)
            self.processor_thread.finished_processing.connect(self.on_processing_finished)
            self.processor_thread.error_occurred.connect(self.on_processing_error)

            # Thread başlat
            self.processor_thread.start()

        except Exception as e:
            self.handle_critical_error("Excel işleme başlatma hatası", str(e))

    def _cleanup_thread(self):
        """Önceki thread'i temizle"""
        if hasattr(self, 'processor_thread') and self.processor_thread:
            try:
                if self.processor_thread.isRunning():
                    self.processor_thread.terminate()
                    self.processor_thread.wait(3000)  # 3 saniye bekle
            except:
                pass
            finally:
                self.processor_thread = None

    def update_progress(self, value):
        """Progress bar'ı güvenli şekilde güncelle"""
        try:
            if hasattr(self, 'progress_bar') and self.progress_bar:
                self.progress_bar.setValue(value)
        except:
            pass

    def add_log_message(self, message):
        """Log mesajı ekle"""
        try:
            if hasattr(self, 'log_text') and self.log_text:
                self.log_text.append(message)
                # Otomatik scroll
                scrollbar = self.log_text.verticalScrollBar()
                scrollbar.setValue(scrollbar.maximum())
        except:
            pass

    def on_processing_finished(self, dayibasi_data, total_amount):
        """İşlem tamamlandığında çağrılır"""
        try:
            self.dayibasi_data = dayibasi_data
            self.total_amount = total_amount

            # UI güncellemeleri
            self.total_dayibasi_label.setText(f"Toplam Dayıbaşı: {len(dayibasi_data)}")
            self.total_amount_label.setText(f"Toplam Tutar: ₺{total_amount:,.2f}")

            # Tabloyu güncelle
            self.update_data_table()

            # Butonları aktif et
            self.process_btn.setEnabled(True)
            self.create_report_btn.setEnabled(len(dayibasi_data) > 0)
            self.progress_bar.setVisible(False)

            # Son mesajlar
            if len(dayibasi_data) > 0:
                self.add_log_message("🎉 İşlem başarıyla tamamlandı!")
                self.add_log_message("📊 Rapor oluşturmaya hazır.")
            else:
                self.add_log_message("⚠️ Geçerli veri bulunamadı.")
                QMessageBox.warning(self, "Sonuç",
                                    "Excel dosyası işlendi ancak geçerli dayıbaşı verisi bulunamadı.\n"
                                    "Dosya formatını kontrol edin.")

        except Exception as e:
            self.handle_critical_error("Sonuç işleme hatası", str(e))

    def on_processing_error(self, error_message):
        """İşlem hatası durumunda çağrılır"""
        try:
            # UI'yi resetle
            self.progress_bar.setVisible(False)
            self.process_btn.setEnabled(True)
            self.create_report_btn.setEnabled(False)

            self.add_log_message(f"❌ İşlem hatası: {error_message}")

            # Kullanıcıya mesaj göster
            if "bulunamadı" in error_message.lower():
                QMessageBox.critical(self, "Dosya Hatası",
                                     f"Dosya problemi:\n{error_message}")
            elif "okunamadı" in error_message.lower():
                QMessageBox.critical(self, "Excel Hatası",
                                     f"Excel okuma hatası:\n{error_message}\n\n"
                                     "Dosyanın Excel formatında ve açık olmadığından emin olun.")
            else:
                QMessageBox.critical(self, "İşlem Hatası", error_message)

        except Exception as e:
            self.handle_critical_error("Hata işleme sırasında hata", str(e))

    def handle_critical_error(self, title, error_msg):
        """Kritik hataları işle"""
        try:
            print(f"CRITICAL ERROR - {title}: {error_msg}")

            # UI'yi güvenli duruma getir
            if hasattr(self, 'progress_bar'):
                self.progress_bar.setVisible(False)
            if hasattr(self, 'process_btn'):
                self.process_btn.setEnabled(True)
            if hasattr(self, 'create_report_btn'):
                self.create_report_btn.setEnabled(False)

            self.add_log_message(f"💥 {title}: {error_msg}")
            QMessageBox.critical(self, title, f"Kritik hata oluştu:\n{error_msg}")

        except:
            print(f"FATAL ERROR - Hata işleyici de çöktü: {error_msg}")

    def update_data_table(self):
        """Veri tablosunu güncelle"""
        try:
            self.data_table.setRowCount(len(self.dayibasi_data))

            for row, data in enumerate(self.dayibasi_data):
                # Güvenli veri çıkarma
                name = data.get('name', '')
                iban = data.get('iban', '')
                bank = data.get('bank', '')
                phone = data.get('phone', '')
                amount = data.get('amount', 0)

                self.data_table.setItem(row, 0, QTableWidgetItem(str(name)))
                self.data_table.setItem(row, 1, QTableWidgetItem(str(iban)))
                self.data_table.setItem(row, 2, QTableWidgetItem(str(bank)))
                self.data_table.setItem(row, 3, QTableWidgetItem(str(phone)))
                self.data_table.setItem(row, 4, QTableWidgetItem(f"₺{float(amount):,.2f}"))

        except Exception as e:
            self.add_log_message(f"⚠️ Tablo güncelleme hatası: {str(e)}")

    def create_payment_report(self):
        """Ödeme raporu oluştur"""
        if not self.dayibasi_data:
            QMessageBox.warning(self, "Uyarı", "Önce Excel dosyasını işleyin.")
            return

        # Kayıt yeri seç
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "Raporu Kaydet",
            f"Odeme_Raporu_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel Dosyaları (*.xlsx);;Tüm Dosyalar (*)"
        )

        if not save_path:
            return

        try:
            # Rapor verilerini hazırla
            start_date_str = self.start_date.date().toString("dd MMMM yyyy")
            end_date_str = self.end_date.date().toString("dd MMMM yyyy")
            region = self.region_combo.currentText()

            # Excel raporu oluştur
            self.create_formatted_excel_report(save_path, start_date_str, end_date_str, region)

            QMessageBox.information(self, "Başarılı", f"Ödeme raporu başarıyla oluşturuldu:\n{save_path}")
            self.add_log_message(f"✅ Rapor oluşturuldu: {os.path.basename(save_path)}")

        except Exception as e:
            QMessageBox.critical(self, "Hata", f"Rapor oluşturulurken hata oluştu:\n{str(e)}")
            self.add_log_message(f"❌ Rapor hatası: {str(e)}")

    def create_formatted_excel_report(self, output_path, start_date, end_date, region):
        """Formatlı Excel raporu oluştur"""
        # Workbook oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = "ÖDEME DETAY LİSTESİ"

        # Stil tanımlamaları
        title_font = Font(name='Arial', size=14, bold=True)
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        data_font = Font(name='Arial', size=11)

        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Başlık bilgileri
        ws['A1'] = "PMI SCV RİMA MANİSA ÖDEME DETAY LİSTESİ"
        ws['A1'].font = title_font
        ws.merge_cells('A1:G1')

        ws['A3'] = f"BÖLGE: {region}"
        ws['A4'] = f"BAŞLANGIÇ TARİHİ: {start_date}"
        ws['A5'] = f"BİTİŞ TARİHİ: {end_date}"

        ws['A7'] = "EFT Masrafları için hesaba yollanacak tutar"
        ws['A8'] = f"HARCAMA TALEP FORMU TUTARI: ₺{self.total_amount:,.2f}"
        ws['A8'].font = Font(name='Arial', size=12, bold=True, color='FF0000')

        ws['A10'] = "ÖDEME DETAYLARI:"
        ws['A10'].font = Font(name='Arial', size=12, bold=True)

        # Tablo başlıkları
        headers = ['BÖLGE', 'KONTRAT', 'AD SOYAD', 'BANKA', 'IBAN', 'ÖDENECEK TUTAR', 'NOT']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=12, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # Veri satırları
        for row_idx, data in enumerate(self.dayibasi_data, 13):
            ws.cell(row=row_idx, column=1, value=region).border = border
            ws.cell(row=row_idx, column=2, value=f"KONTRAT-{row_idx - 12:03d}").border = border
            ws.cell(row=row_idx, column=3, value=data.get('name', '')).border = border
            ws.cell(row=row_idx, column=4, value=data.get('bank', '')).border = border
            ws.cell(row=row_idx, column=5, value=data.get('iban', '')).border = border

            # ÖDENECEK TUTAR
            amount = data.get('amount', 0)
            amount_cell = ws.cell(row=row_idx, column=6, value=f"₺{float(amount):,.2f}")
            amount_cell.border = border
            amount_cell.alignment = Alignment(horizontal='right', vertical='center')

            ws.cell(row=row_idx, column=7, value='').border = border

        # Sütun genişliklerini ayarla
        column_widths = [10, 15, 25, 15, 30, 18, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

        # Dosyayı kaydet
        wb.save(output_path)

    def closeEvent(self, event):
        """Pencere kapatılırken temizlik yap"""
        try:
            self._cleanup_thread()
            super().closeEvent(event)
        except:
            super().closeEvent(event)

    def update_data(self, data=None):
        """Veri güncellendiğinde çağrılır - Diğer tablarla uyumluluk için"""
        pass